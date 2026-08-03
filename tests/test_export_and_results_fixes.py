"""Mobile-first Excel sizing + partial-results stats regression tests."""
from openpyxl import Workbook

from services.excel_mobile import MAX_COL_WIDTH, mobile_optimize


def test_mobile_optimize_clamps_wide_columns_and_wraps():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'a very long answer that used to need a desktop-width column'
    ws['B1'] = 'narrow'
    ws.column_dimensions['A'].width = 90
    ws.column_dimensions['B'].width = 20

    mobile_optimize(wb)

    assert ws.column_dimensions['A'].width == MAX_COL_WIDTH
    assert ws.column_dimensions['B'].width == 20  # untouched
    assert ws['A1'].alignment.wrap_text is True
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.orientation == 'portrait'


def test_all_generators_apply_mobile_optimize():
    # Source-level guard: every Excel generator must call mobile_optimize.
    import inspect
    from services import bestprep_excel, excel_dashboard
    from services.smart_analysis import excel_generator as smart_excel
    from services.scraper.agents.presentation import excel_generator as scraper_excel

    for module in (excel_dashboard, bestprep_excel, smart_excel, scraper_excel):
        assert 'mobile_optimize' in inspect.getsource(module), module.__name__


def test_partial_browser_output_includes_stats():
    """Regression: mid-packaging/stopped fetches rendered '0/0 questions, 0 pages'
    because the partial output lacked the top-level stat keys."""
    from services.hotdog.models import ParsedConfig, Question, Section
    from services.hotdog.orchestrator import HotdogOrchestrator

    orch = HotdogOrchestrator(openai_api_key='sk-test-offline')
    question = Question(id='Q1', text='What is the scope?', section_id='s1')
    section = Section(id='s1', name='Scope', description='', questions=[question])
    config = ParsedConfig(
        name='test',
        version='1.0',
        sections=[section],
        section_map={'s1': section},
        question_map={'Q1': question},
    )

    out = orch._build_partial_browser_output({}, config)
    assert out['total_questions'] == 1
    assert out['questions_answered'] == 0
    assert 'total_pages' in out
    assert out['sections'][0]['questions'][0]['has_answer'] is False


# ---- Wrapped answers must actually be visible (2.5.2) ------------------------
# The mobile clamp narrows the Answer column to 42 characters and turns on
# wrapping, but the generator set a hard 55pt row height. A long answer wrapped
# into ~15 lines and everything past line four was invisible in the workbook.

def _dashboard_with_answer(answer):
    from services.excel_dashboard import ExcelDashboardGenerator
    result = {'sections': [{'section_name': 'Scope', 'questions': [
        {'question': 'Scope?', 'answer': answer + ' <PDF pg 12>', 'page_citations': [12],
         'answer_summary': 'Distilled.', 'footnote': 'x'},
        {'question': 'Short?', 'answer': 'Yes <PDF pg 2>', 'page_citations': [2]},
    ]}]}
    generator = ExcelDashboardGenerator(result)
    generator.generate()
    return generator.wb['Detailed Results']


def _long_text(n):
    chunk = ('The contractor shall furnish all labor and materials for the complete '
             'CIPP lining installation per ASTM F1216 including bypass pumping. ')
    return (chunk * (n // len(chunk) + 1))[:n]


def test_a_long_answer_row_grows_beyond_the_hardcoded_height():
    ws = _dashboard_with_answer(_long_text(800))
    height = ws.row_dimensions[3].height
    assert height > 55, 'the fixed 55pt height is what hid wrapped answers'
    width = ws.column_dimensions['D'].width
    lines = 800 / (width - 1)
    assert height >= lines * 14, f'{height}pt cannot show ~{lines:.0f} wrapped lines'


def test_a_short_answer_row_is_not_inflated():
    ws = _dashboard_with_answer('Yes')
    assert ws.row_dimensions[4].height <= 60, 'short rows must stay compact'


def test_normal_content_keeps_the_phone_friendly_clamp():
    ws = _dashboard_with_answer(_long_text(300))
    assert ws.column_dimensions['D'].width == 42, \
        'ordinary answers must not widen columns - the export is phone-first'


def test_a_column_widens_only_when_the_clamp_would_clip_it():
    """Excel caps a row at 409.5pt. If 42 characters would push the text past
    that, the column widens just enough rather than silently hiding the tail."""
    ws = _dashboard_with_answer(_long_text(1400))
    width = ws.column_dimensions['D'].width
    assert width > 42, 'a column that would clip must widen'
    assert width <= 72, 'widening is bounded - it must not become a desktop-only sheet'


def test_row_height_never_exceeds_the_excel_ceiling():
    ws = _dashboard_with_answer(_long_text(6000))
    assert ws.row_dimensions[3].height <= 409.5, 'Excel rejects taller rows'


def test_header_rows_keep_their_deliberate_heights():
    ws = _dashboard_with_answer(_long_text(800))
    assert ws.row_dimensions[2].height >= 25, 'the header band must not be shrunk'
