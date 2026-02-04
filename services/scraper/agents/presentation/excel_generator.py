"""
Excel Generator Agent (PR-2)

Creates styled Excel workbooks from extraction results using openpyxl.

Key Responsibilities:
- Create Excel workbook with multiple sheets
- Style headers with colors and borders
- Alternating row colors for readability
- Auto-column width adjustment
- Freeze header rows
- Sheet for Systems Info data
- Sheet for Public Bids data
- Sheet for Sources summary
- No truncation - show all data

Version: 1.0.0
"""

import logging
import os
import re
import time
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.scraper.agents.base import BaseAgent
from services.scraper.models import (
    AgentRequest,
    AgentResponse,
    ExtractionResult,
    MunicipalSystemsInfoRow,
    MunicipalPublicBidRow,
    Municipality,
    ExtractedDataPoint,
    SourceURL,
    VerbatimCitation,
)
from services.scraper.prompts.pr2_excel_generator import (
    get_config,
    SYSTEMS_INFO_SHEET,
    PUBLIC_BIDS_SHEET,
    SOURCES_SHEET,
    SUMMARY_SHEET,
)

logger = logging.getLogger(__name__)


def sanitize_for_excel(text: Any) -> str:
    """
    Remove or replace characters that are illegal in Excel worksheet cells.
    Excel doesn't allow control characters (0x00-0x1F except tab, newline, carriage return).
    """
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)

    # Remove illegal XML characters (control chars except tab, newline, carriage return)
    illegal_pattern = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
    cleaned = illegal_pattern.sub('', text)

    # Replace unit separator with space
    cleaned = cleaned.replace('\u001f', ' ')

    return cleaned


class ExcelGeneratorAgent(BaseAgent):
    """
    PR-2: Excel Generator Agent

    Creates styled Excel workbooks from extraction results.
    Multiple sheets: Systems Info, Public Bids, Sources, Summary.
    Professional styling with headers, borders, alternating rows.

    This agent does NOT need LLM - it performs mechanical Excel generation.
    """

    AGENT_ID = "pr-2"
    AGENT_NAME = "Excel Generator"
    AGENT_VERSION = "1.0.0"
    PROMPT_VERSION = "1.0.0"  # Config version, not LLM prompt
    PROMPT_LAST_REFINED = "2026-02-04"

    # BidBrief brand colors
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
    ALT_ROW_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # Light blue
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Additional styling
    TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    TITLE_FONT = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    DATA_FONT = Font(name='Calibri', size=10, color="000000")
    DATA_FONT_BOLD = Font(name='Calibri', size=10, bold=True, color="000000")
    LINK_FONT = Font(name='Calibri', size=10, color="0563C1", underline='single')

    def __init__(self, *args, **kwargs):
        """Initialize the Excel Generator Agent."""
        super().__init__(*args, **kwargs)
        self._config = get_config()

    def get_system_prompt(self) -> str:
        """
        Get system prompt - not used for this mechanical agent.

        The Excel Generator does purely mechanical formatting, so no LLM
        calls are needed for standard operation.
        """
        return """You are an Excel Generator specialist for municipal data.

This agent performs mechanical Excel generation and does not typically
require LLM assistance. If invoked, help with formatting decisions."""

    def _create_workbook(self) -> Workbook:
        """Create new workbook with default styling."""
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        return wb

    def _get_safe_value(self, value: Any) -> str:
        """
        Safely extract string value from various types.

        Args:
            value: The value to convert (str, ExtractedDataPoint, list, dict, etc.)

        Returns:
            Sanitized string suitable for Excel cell
        """
        if value is None:
            return "NOT FOUND"

        if isinstance(value, ExtractedDataPoint):
            return sanitize_for_excel(value.value or "NOT FOUND")

        if isinstance(value, list):
            if not value:
                return "NOT FOUND"
            items = [str(item) for item in value if item]
            return sanitize_for_excel(", ".join(items))

        if isinstance(value, dict):
            if not value:
                return "NOT FOUND"
            parts = [f"{k}: {v}" for k, v in value.items() if v]
            return sanitize_for_excel("; ".join(parts))

        text = str(value).strip()
        if not text:
            return "NOT FOUND"
        return sanitize_for_excel(text)

    def _format_sources(self, source_urls: List[Any]) -> str:
        """Format source URLs for Excel cell."""
        if not source_urls:
            return "NOT FOUND"

        urls = []
        for source in source_urls:
            if isinstance(source, SourceURL):
                urls.append(source.url)
            elif hasattr(source, 'url'):
                urls.append(source.url)
            elif isinstance(source, dict) and 'url' in source:
                urls.append(source['url'])
            elif isinstance(source, str):
                urls.append(source)

        if not urls:
            return "NOT FOUND"

        return sanitize_for_excel("\n".join(urls))

    def _format_citations(self, verbatim_citations: List[Any]) -> str:
        """
        Format verbatim citations for Excel cell.

        NO TRUNCATION - preserve full citations.
        """
        if not verbatim_citations:
            return "NOT FOUND"

        texts = []
        for citation in verbatim_citations:
            if isinstance(citation, VerbatimCitation):
                texts.append(citation.text)
            elif hasattr(citation, 'text'):
                texts.append(citation.text)
            elif isinstance(citation, dict) and 'text' in citation:
                texts.append(citation['text'])
            elif isinstance(citation, str):
                texts.append(citation)

        if not texts:
            return "NOT FOUND"

        # Join with double newline for readability - NO TRUNCATION
        return sanitize_for_excel("\n\n".join(texts))

    def _style_header_row(self, ws, row_num: int):
        """Apply header styling to a row."""
        for cell in ws[row_num]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.THIN_BORDER
        ws.row_dimensions[row_num].height = 30

    def _style_data_rows(self, ws, start_row: int, end_row: int):
        """Apply alternating row colors and borders."""
        for row_idx in range(start_row, end_row + 1):
            is_alt = (row_idx - start_row) % 2 == 1
            for cell in ws[row_idx]:
                cell.font = self.DATA_FONT
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = self.THIN_BORDER
                if is_alt:
                    cell.fill = self.ALT_ROW_FILL
            ws.row_dimensions[row_idx].height = 40

    def _auto_column_width(self, ws, predefined_widths: Optional[Dict[str, int]] = None):
        """
        Adjust column widths based on content.

        Args:
            ws: Worksheet to adjust
            predefined_widths: Optional dict mapping column index (1-based) or header to width
        """
        config = self._config['column_widths']
        min_width = config['min']
        max_width = config['max']
        default_width = config['default']

        for column_cells in ws.columns:
            col_idx = column_cells[0].column
            col_letter = get_column_letter(col_idx)

            # Check if predefined width exists
            if predefined_widths and col_idx in predefined_widths:
                ws.column_dimensions[col_letter].width = predefined_widths[col_idx]
                continue

            # Calculate width based on content
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        # For multiline cells, find the longest line
                        cell_text = str(cell.value)
                        lines = cell_text.split('\n')
                        for line in lines:
                            max_length = max(max_length, len(line))
                except (AttributeError, TypeError, ValueError):
                    # Skip cells that can't be processed for width calculation
                    pass

            # Apply width with limits
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _freeze_header(self, ws):
        """Freeze the header row."""
        ws.freeze_panes = 'A2'

    def _add_summary_sheet(self, wb: Workbook, result: ExtractionResult, municipality: Municipality):
        """Add Summary sheet with extraction overview."""
        ws = wb.create_sheet(SUMMARY_SHEET, 0)

        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = 'CITYSCRAPER EXTRACTION SUMMARY'
        title_cell.font = self.TITLE_FONT
        title_cell.fill = self.TITLE_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        row = 3

        # Municipality info
        summary_data = [
            ('Municipality:', municipality.full_name),
            ('State:', municipality.state),
            ('County:', municipality.county or 'N/A'),
            ('', ''),  # Spacer
            ('Extraction Statistics', ''),
            ('Total Sources Searched:', str(result.total_sources_searched)),
            ('Total Data Points Extracted:', str(result.total_data_points_extracted)),
            ('Systems Info Rows:', str(len(result.systems_info_rows))),
            ('Public Bid Rows:', str(len(result.public_bid_rows))),
            ('Downloaded Documents:', str(len(result.downloaded_documents))),
            ('', ''),  # Spacer
            ('Timing', ''),
            ('Started At:', result.started_at.strftime('%Y-%m-%d %H:%M:%S') if result.started_at else 'N/A'),
            ('Completed At:', result.completed_at.strftime('%Y-%m-%d %H:%M:%S') if result.completed_at else 'N/A'),
            ('', ''),  # Spacer
            ('Data Quality', ''),
            ('Data Gaps:', str(len(result.data_gaps))),
            ('Conflicts Detected:', str(len(result.conflicts_detected))),
        ]

        for label, value in summary_data:
            if label == '' and value == '':
                row += 1
                continue

            if value == '':
                # Section header
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="1F4E79")
                ws[f'A{row}'].fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            else:
                ws.cell(row, 1, label).font = Font(name='Calibri', size=11, bold=True)
                ws.cell(row, 2, value).font = self.DATA_FONT

            for col in range(1, 3):
                ws.cell(row, col).border = self.THIN_BORDER
                ws.cell(row, col).alignment = Alignment(vertical='center')

            ws.row_dimensions[row].height = 22
            row += 1

        # Data gaps if any
        if result.data_gaps:
            row += 1
            ws[f'A{row}'] = 'Data Gaps Details:'
            ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            row += 1
            for gap in result.data_gaps:
                ws[f'A{row}'] = f"  - {gap}"
                ws[f'A{row}'].font = self.DATA_FONT
                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

    def _add_systems_info_sheet(self, wb: Workbook, rows: List[MunicipalSystemsInfoRow]):
        """Add Municipal Systems Information sheet."""
        ws = wb.create_sheet(SYSTEMS_INFO_SHEET)

        # Define headers matching extractiondev.md
        headers = [
            "Municipality (City)",
            "State",
            "Relevant Agency",
            "Agency & Scope",
            "Sanitary Sewer Pipe",
            "Storm Drain Pipe",
            "Storm Drain Assets",
            "System Age/History",
            "Owned Equipment",
            "Maintenance Practices",
            "Sewage Incidents",
            "Storm Incidents",
            "Source URLs",
            "Verbatim Citations",
            "Notes/Reconciliation"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)

        self._style_header_row(ws, 1)
        self._freeze_header(ws)

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            ws.cell(row_idx, 1, self._get_safe_value(row_data.municipality_city))
            ws.cell(row_idx, 2, self._get_safe_value(row_data.state))
            ws.cell(row_idx, 3, self._get_safe_value(row_data.relevant_agency))
            ws.cell(row_idx, 4, self._get_safe_value(row_data.agency_scope))
            ws.cell(row_idx, 5, self._get_safe_value(row_data.sanitary_sewer_pipe))
            ws.cell(row_idx, 6, self._get_safe_value(row_data.storm_drain_pipe))
            ws.cell(row_idx, 7, self._get_safe_value(row_data.storm_drain_assets))
            ws.cell(row_idx, 8, self._get_safe_value(row_data.system_age_history))
            ws.cell(row_idx, 9, self._get_safe_value(row_data.equipment_owned))
            ws.cell(row_idx, 10, self._get_safe_value(row_data.maintenance_practices))
            ws.cell(row_idx, 11, self._get_safe_value(row_data.sewage_incidents))
            ws.cell(row_idx, 12, self._get_safe_value(row_data.storm_incidents))
            ws.cell(row_idx, 13, self._format_sources(row_data.source_urls))
            ws.cell(row_idx, 14, self._format_citations(row_data.verbatim_citations))
            ws.cell(row_idx, 15, self._get_safe_value(row_data.notes_reconciliation))

        # Style data rows
        if rows:
            self._style_data_rows(ws, 2, len(rows) + 1)

        # Auto-adjust column widths
        predefined = {
            1: 20, 2: 8, 3: 25, 4: 40, 5: 45, 6: 45, 7: 40, 8: 40,
            9: 40, 10: 45, 11: 40, 12: 40, 13: 50, 14: 60, 15: 40
        }
        self._auto_column_width(ws, predefined)

    def _add_public_bids_sheet(self, wb: Workbook, rows: List[MunicipalPublicBidRow]):
        """Add Municipal Public Bids sheet."""
        ws = wb.create_sheet(PUBLIC_BIDS_SHEET)

        # Define headers matching extractiondev.md
        headers = [
            "Municipality (City)",
            "State",
            "Agency/Municipality",
            "Bid/Contract Title",
            "Sewer/Storm Keywords",
            "Scope",
            "Timeline & Requirements",
            "Contacts",
            "Status + Key Dates",
            "Source URLs",
            "Verbatim Citations",
            "Notes/Reconciliation"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)

        self._style_header_row(ws, 1)
        self._freeze_header(ws)

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            # Format status with key dates
            status_parts = [row_data.status or "NOT FOUND"]
            if row_data.key_dates:
                date_parts = [f"{k}: {v}" for k, v in row_data.key_dates.items() if v]
                if date_parts:
                    status_parts.append("; ".join(date_parts))
            status_with_dates = " | ".join(status_parts)

            # Format keywords
            keywords = ", ".join(row_data.sewer_storm_keywords) if row_data.sewer_storm_keywords else "NOT FOUND"

            ws.cell(row_idx, 1, self._get_safe_value(row_data.municipality_city))
            ws.cell(row_idx, 2, self._get_safe_value(row_data.state))
            ws.cell(row_idx, 3, self._get_safe_value(row_data.agency_municipality))
            ws.cell(row_idx, 4, self._get_safe_value(row_data.bid_contract_title))
            ws.cell(row_idx, 5, sanitize_for_excel(keywords))
            ws.cell(row_idx, 6, self._get_safe_value(row_data.scope))
            ws.cell(row_idx, 7, self._get_safe_value(row_data.timeline_requirements))
            ws.cell(row_idx, 8, self._get_safe_value(row_data.contacts))
            ws.cell(row_idx, 9, sanitize_for_excel(status_with_dates))
            ws.cell(row_idx, 10, self._format_sources(row_data.source_urls))
            ws.cell(row_idx, 11, self._format_citations(row_data.verbatim_citations))
            ws.cell(row_idx, 12, self._get_safe_value(row_data.notes_reconciliation))

            # Apply status-based styling to status cell
            status_cell = ws.cell(row_idx, 9)
            status_lower = (row_data.status or "").lower()
            if status_lower == "open":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status_lower == "closed":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif status_lower == "awarded":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Style data rows
        if rows:
            self._style_data_rows(ws, 2, len(rows) + 1)

        # Auto-adjust column widths
        predefined = {
            1: 20, 2: 8, 3: 25, 4: 40, 5: 30, 6: 50,
            7: 40, 8: 35, 9: 30, 10: 50, 11: 60, 12: 40
        }
        self._auto_column_width(ws, predefined)

    def _add_sources_sheet(self, wb: Workbook, result: ExtractionResult):
        """Add Sources summary sheet with all URLs and citations."""
        ws = wb.create_sheet(SOURCES_SHEET)

        # Define headers
        headers = ["#", "URL", "Title", "Source Type", "Retrieved At", "Relevance"]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)

        self._style_header_row(ws, 1)
        self._freeze_header(ws)

        # Collect all sources from systems info and public bids
        all_sources = []

        for row_data in result.systems_info_rows:
            for source in row_data.source_urls:
                if isinstance(source, SourceURL):
                    all_sources.append({
                        'url': source.url,
                        'title': source.title,
                        'source_type': source.source_type,
                        'retrieved_at': source.retrieved_at.strftime('%Y-%m-%d %H:%M') if source.retrieved_at else 'N/A',
                        'relevance': f"{source.relevance_score:.2f}" if source.relevance_score else 'N/A',
                    })
                elif isinstance(source, dict):
                    all_sources.append({
                        'url': source.get('url', 'N/A'),
                        'title': source.get('title', 'N/A'),
                        'source_type': source.get('source_type', 'N/A'),
                        'retrieved_at': source.get('retrieved_at', 'N/A'),
                        'relevance': source.get('relevance_score', 'N/A'),
                    })
                elif isinstance(source, str):
                    all_sources.append({
                        'url': source,
                        'title': 'N/A',
                        'source_type': 'N/A',
                        'retrieved_at': 'N/A',
                        'relevance': 'N/A',
                    })

        for row_data in result.public_bid_rows:
            for source in row_data.source_urls:
                if isinstance(source, SourceURL):
                    all_sources.append({
                        'url': source.url,
                        'title': source.title,
                        'source_type': source.source_type,
                        'retrieved_at': source.retrieved_at.strftime('%Y-%m-%d %H:%M') if source.retrieved_at else 'N/A',
                        'relevance': f"{source.relevance_score:.2f}" if source.relevance_score else 'N/A',
                    })
                elif isinstance(source, dict):
                    all_sources.append({
                        'url': source.get('url', 'N/A'),
                        'title': source.get('title', 'N/A'),
                        'source_type': source.get('source_type', 'N/A'),
                        'retrieved_at': source.get('retrieved_at', 'N/A'),
                        'relevance': source.get('relevance_score', 'N/A'),
                    })
                elif isinstance(source, str):
                    all_sources.append({
                        'url': source,
                        'title': 'N/A',
                        'source_type': 'N/A',
                        'retrieved_at': 'N/A',
                        'relevance': 'N/A',
                    })

        # Remove duplicates by URL
        seen_urls = set()
        unique_sources = []
        for source in all_sources:
            if source['url'] not in seen_urls:
                seen_urls.add(source['url'])
                unique_sources.append(source)

        # Write data rows
        for row_idx, source in enumerate(unique_sources, 2):
            ws.cell(row_idx, 1, row_idx - 1)  # Source number
            url_cell = ws.cell(row_idx, 2, sanitize_for_excel(source['url']))
            url_cell.font = self.LINK_FONT
            ws.cell(row_idx, 3, sanitize_for_excel(source['title']))
            ws.cell(row_idx, 4, sanitize_for_excel(source['source_type']))
            ws.cell(row_idx, 5, sanitize_for_excel(str(source['retrieved_at'])))
            ws.cell(row_idx, 6, sanitize_for_excel(str(source['relevance'])))

        # Style data rows
        if unique_sources:
            self._style_data_rows(ws, 2, len(unique_sources) + 1)

        # Auto-adjust column widths
        predefined = {1: 8, 2: 60, 3: 40, 4: 15, 5: 20, 6: 15}
        self._auto_column_width(ws, predefined)

        # Add "No sources found" message if empty
        if not unique_sources:
            ws.cell(2, 1, "No sources collected during extraction")
            ws.merge_cells('A2:F2')
            ws['A2'].font = Font(name='Calibri', size=11, italic=True, color="666666")

    async def process(self, request: AgentRequest) -> AgentResponse:
        """
        Process Excel generation request.

        Expected input_data:
        - extraction_result: ExtractionResult object (or dict representation)
        - municipality: Municipality object (or dict representation)
        - filename: str (optional, default generated)
        - output_dir: str (optional, default 'exports')

        Returns AgentResponse with:
        - file_path: Path to generated Excel file
        - file_size_bytes: Size of the file
        - sheets_created: List of sheet names
        - row_counts: Dict of row counts per sheet
        """
        start_time = time.time()

        try:
            self.emit_event("processing", "Generating Excel workbook")

            # Extract inputs
            extraction_result = request.input_data.get('extraction_result')
            municipality_data = request.input_data.get('municipality')
            filename = request.input_data.get('filename')
            output_dir = request.input_data.get('output_dir', 'exports')

            # Validate extraction_result
            if extraction_result is None:
                return AgentResponse(
                    agent_id=self.AGENT_ID,
                    task=request.task,
                    success=False,
                    output_data={},
                    errors=["extraction_result is required"],
                    processing_time_seconds=time.time() - start_time
                )

            # Convert extraction_result if dict
            if isinstance(extraction_result, dict):
                # Build ExtractionResult from dict
                extraction_result = self._dict_to_extraction_result(extraction_result)
            elif not isinstance(extraction_result, ExtractionResult):
                return AgentResponse(
                    agent_id=self.AGENT_ID,
                    task=request.task,
                    success=False,
                    output_data={},
                    errors=[f"extraction_result must be ExtractionResult or dict, got {type(extraction_result).__name__}"],
                    processing_time_seconds=time.time() - start_time
                )

            # Get municipality
            if municipality_data is None:
                municipality = extraction_result.municipality
            elif isinstance(municipality_data, dict):
                municipality = Municipality(
                    city=municipality_data.get('city', 'Unknown'),
                    state=municipality_data.get('state', 'Unknown'),
                    county=municipality_data.get('county'),
                    region=municipality_data.get('region'),
                )
            elif isinstance(municipality_data, Municipality):
                municipality = municipality_data
            else:
                municipality = extraction_result.municipality

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                safe_city = re.sub(r'[^\w\s-]', '', municipality.city).replace(' ', '_')
                filename = f"CityScraper_{safe_city}_{municipality.state}_{timestamp}.xlsx"

            # Ensure output directory exists
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            file_path = os.path.join(output_dir, filename)

            # Create workbook
            self.emit_event("processing", "Creating workbook structure")
            wb = self._create_workbook()

            sheets_created = []
            row_counts = {}

            # Add Summary sheet
            self._add_summary_sheet(wb, extraction_result, municipality)
            sheets_created.append(SUMMARY_SHEET)
            row_counts[SUMMARY_SHEET] = 1

            # Add Systems Info sheet if data exists
            if extraction_result.systems_info_rows:
                self.emit_event("processing", f"Adding {len(extraction_result.systems_info_rows)} systems info rows")
                self._add_systems_info_sheet(wb, extraction_result.systems_info_rows)
                sheets_created.append(SYSTEMS_INFO_SHEET)
                row_counts[SYSTEMS_INFO_SHEET] = len(extraction_result.systems_info_rows)

            # Add Public Bids sheet if data exists
            if extraction_result.public_bid_rows:
                self.emit_event("processing", f"Adding {len(extraction_result.public_bid_rows)} public bid rows")
                self._add_public_bids_sheet(wb, extraction_result.public_bid_rows)
                sheets_created.append(PUBLIC_BIDS_SHEET)
                row_counts[PUBLIC_BIDS_SHEET] = len(extraction_result.public_bid_rows)

            # Add Sources sheet
            self._add_sources_sheet(wb, extraction_result)
            sheets_created.append(SOURCES_SHEET)

            # Count unique sources for row count
            all_source_urls = set()
            for row in extraction_result.systems_info_rows:
                for source in row.source_urls:
                    if isinstance(source, SourceURL):
                        all_source_urls.add(source.url)
                    elif isinstance(source, dict):
                        all_source_urls.add(source.get('url', ''))
                    elif isinstance(source, str):
                        all_source_urls.add(source)
            for row in extraction_result.public_bid_rows:
                for source in row.source_urls:
                    if isinstance(source, SourceURL):
                        all_source_urls.add(source.url)
                    elif isinstance(source, dict):
                        all_source_urls.add(source.get('url', ''))
                    elif isinstance(source, str):
                        all_source_urls.add(source)
            row_counts[SOURCES_SHEET] = len(all_source_urls)

            # Save workbook
            self.emit_event("processing", "Saving workbook")
            wb.save(file_path)

            # Get file size
            file_size_bytes = os.path.getsize(file_path)

            elapsed = time.time() - start_time

            self.emit_event(
                "completed",
                f"Generated Excel with {len(sheets_created)} sheets, {sum(row_counts.values())} total rows",
                is_completed=True
            )

            return AgentResponse(
                agent_id=self.AGENT_ID,
                task=request.task,
                success=True,
                output_data={
                    'file_path': file_path,
                    'file_size_bytes': file_size_bytes,
                    'sheets_created': sheets_created,
                    'row_counts': row_counts,
                },
                processing_time_seconds=elapsed
            )

        except Exception as e:
            logger.exception(f"PR-2 unexpected error: {e}")
            self.emit_event("failed", f"Excel generation error: {str(e)[:100]}")
            return AgentResponse(
                agent_id=self.AGENT_ID,
                task=request.task,
                success=False,
                output_data={},
                errors=[f"Unexpected error: {type(e).__name__}: {str(e)[:200]}"],
                processing_time_seconds=time.time() - start_time
            )

    def _dict_to_extraction_result(self, data: dict) -> ExtractionResult:
        """Convert a dict representation to ExtractionResult."""
        from services.scraper.models import (
            Municipality, TableMode, PreflightResult, PreflightStatus
        )

        # Get municipality
        muni_data = data.get('municipality', {})
        if isinstance(muni_data, Municipality):
            municipality = muni_data
        else:
            municipality = Municipality(
                city=muni_data.get('city', 'Unknown'),
                state=muni_data.get('state', 'Unknown'),
                county=muni_data.get('county'),
                region=muni_data.get('region'),
            )

        # Get table mode
        mode_val = data.get('table_mode', TableMode.MUNICIPAL_SYSTEMS_INFO)
        if isinstance(mode_val, str):
            try:
                table_mode = TableMode(mode_val)
            except ValueError:
                table_mode = TableMode.MUNICIPAL_SYSTEMS_INFO
        elif isinstance(mode_val, TableMode):
            table_mode = mode_val
        else:
            table_mode = TableMode.MUNICIPAL_SYSTEMS_INFO

        # Get preflight
        preflight_data = data.get('preflight', {})
        if isinstance(preflight_data, PreflightResult):
            preflight = preflight_data
        else:
            preflight = PreflightResult(
                municipality=municipality,
                table_mode=table_mode,
                status=PreflightStatus.PASS,
            )

        # Build result
        result = ExtractionResult(
            municipality=municipality,
            table_mode=table_mode,
            preflight=preflight,
            systems_info_rows=data.get('systems_info_rows', []),
            public_bid_rows=data.get('public_bid_rows', []),
            total_sources_searched=data.get('total_sources_searched', 0),
            total_data_points_extracted=data.get('total_data_points_extracted', 0),
            data_gaps=data.get('data_gaps', []),
            conflicts_detected=data.get('conflicts_detected', []),
            downloaded_documents=data.get('downloaded_documents', []),
        )

        if data.get('started_at'):
            result.started_at = data['started_at']
        if data.get('completed_at'):
            result.completed_at = data['completed_at']

        return result

    def validate_output(self, output: Dict[str, Any]) -> List[str]:
        """Validate Excel generator output."""
        errors = []

        # Check required fields
        if 'file_path' not in output:
            errors.append("Missing 'file_path' in output")

        if 'file_size_bytes' not in output:
            errors.append("Missing 'file_size_bytes' in output")

        if 'sheets_created' not in output:
            errors.append("Missing 'sheets_created' in output")

        if 'row_counts' not in output:
            errors.append("Missing 'row_counts' in output")

        # Validate file_path
        if 'file_path' in output:
            file_path = output['file_path']
            if not isinstance(file_path, str):
                errors.append("'file_path' must be a string")
            elif not os.path.exists(file_path):
                errors.append(f"Generated file does not exist: {file_path}")

        # Validate file_size_bytes is positive
        if 'file_size_bytes' in output:
            if not isinstance(output['file_size_bytes'], int):
                errors.append("'file_size_bytes' must be an integer")
            elif output['file_size_bytes'] <= 0:
                errors.append("'file_size_bytes' must be positive")

        # Validate sheets_created is a list
        if 'sheets_created' in output and not isinstance(output['sheets_created'], list):
            errors.append("'sheets_created' must be a list")

        # Validate row_counts is a dict
        if 'row_counts' in output and not isinstance(output['row_counts'], dict):
            errors.append("'row_counts' must be a dictionary")

        return errors
