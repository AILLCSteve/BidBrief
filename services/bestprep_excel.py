"""
BestPrep Excel Export - Comprehensive answer format with all fragments and footnotes.
"""
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_for_excel(text):
    """
    Remove or replace characters that are illegal in Excel worksheet cells.
    Excel doesn't allow control characters (0x00-0x1F except tab, newline, carriage return).
    """
    if not text:
        return text
    if not isinstance(text, str):
        return str(text)

    # Remove illegal XML characters (control chars except tab, newline, carriage return)
    # These cause "cannot be used in worksheets" errors
    illegal_pattern = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
    cleaned = illegal_pattern.sub('', text)

    # Also replace other problematic characters
    # \u001f is Unit Separator - replace with space
    cleaned = cleaned.replace('\u001f', ' ')

    return cleaned


class BestPrepExcelGenerator:
    """Generate exhaustive Excel report for BestPrep mode."""

    # Colors
    NAVY = "104090"
    BLUE = "5E86D0"
    GREEN = "22C55E"
    GRAY = "F3F4F6"
    LIGHT_GRAY = "F9FAFB"
    ORANGE = "FF9800"
    LIGHT_ORANGE = "FFF3E0"

    def __init__(self, analysis_result: dict, accumulator_data: dict):
        self.result = analysis_result
        self.accumulator = accumulator_data
        self.wb = Workbook()
        # Define common styles
        self.thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.label_font = Font(bold=True, color=self.NAVY, size=11)
        self.value_font = Font(color="333333", size=11)

    def generate(self) -> io.BytesIO:
        """Generate 5-sheet BestPrep report."""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self._create_summary_sheet()      # Sheet 1: Overview
        self._create_answers_sheet()      # Sheet 2: Synthesized Answers
        self._create_fragments_sheet()    # Sheet 3: All Fragments
        self._create_footnotes_sheet()    # Sheet 4: All Footnotes
        self._create_sources_sheet()      # Sheet 5: Page Index

        from services.excel_mobile import mobile_optimize
        mobile_optimize(self.wb)

        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _create_summary_sheet(self):
        """Sheet 1: Analysis summary and statistics with professional styling."""
        ws = self.wb.create_sheet("Summary", 0)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50

        stats = self.accumulator.get('statistics', {})

        # Title row
        ws.cell(1, 1, "BestPrep Analysis Summary")
        ws.cell(1, 1).font = Font(size=18, bold=True, color=self.NAVY)
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.cell(2, 1, "Comprehensive Document Analysis Report")
        ws.cell(2, 1).font = Font(size=11, italic=True, color="666666")
        ws.merge_cells('A2:B2')

        # Analysis Info section
        row = 4
        ws.cell(row, 1, "ANALYSIS INFO")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.NAVY)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=self.NAVY)
        ws.merge_cells(f'A{row}:B{row}')

        info_data = [
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Analysis Mode", "BestPrep (Exhaustive)"),
            ("Windows Processed", stats.get('windows_processed', 0)),
        ]

        for label, value in info_data:
            row += 1
            ws.cell(row, 1, label)
            ws.cell(row, 1).font = self.label_font
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.LIGHT_GRAY)
            ws.cell(row, 1).border = self.thin_border
            ws.cell(row, 2, value)
            ws.cell(row, 2).font = self.value_font
            ws.cell(row, 2).border = self.thin_border

        # Question Statistics section
        row += 2
        ws.cell(row, 1, "QUESTION STATISTICS")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.BLUE)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=self.BLUE)
        ws.merge_cells(f'A{row}:B{row}')

        question_data = [
            ("Total Questions", stats.get('total_questions', 0)),
            ("Questions Answered", stats.get('questions_with_answers', 0)),
            ("Questions Synthesized", stats.get('questions_synthesized', 0)),
        ]

        for label, value in question_data:
            row += 1
            ws.cell(row, 1, label)
            ws.cell(row, 1).font = self.label_font
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.LIGHT_GRAY)
            ws.cell(row, 1).border = self.thin_border
            ws.cell(row, 2, value)
            ws.cell(row, 2).font = self.value_font
            ws.cell(row, 2).border = self.thin_border

        # Fragment & Citation Statistics section
        row += 2
        ws.cell(row, 1, "FRAGMENT & CITATION STATISTICS")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.ORANGE)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=self.ORANGE)
        ws.merge_cells(f'A{row}:B{row}')

        fragment_data = [
            ("Total Fragments Collected", stats.get('total_fragments', 0)),
            ("Total Footnotes Extracted", stats.get('total_footnotes', 0)),
            ("Avg Fragments/Question", f"{stats.get('avg_fragments_per_question', 0):.1f}"),
            ("Avg Footnotes/Question", f"{stats.get('avg_footnotes_per_question', 0):.1f}"),
        ]

        for label, value in fragment_data:
            row += 1
            ws.cell(row, 1, label)
            ws.cell(row, 1).font = Font(bold=True, color=self.ORANGE, size=11)
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.LIGHT_ORANGE)
            ws.cell(row, 1).border = self.thin_border
            ws.cell(row, 2, value)
            ws.cell(row, 2).font = Font(bold=True, color=self.ORANGE, size=11)
            ws.cell(row, 2).border = self.thin_border

    def _create_answers_sheet(self):
        """Sheet 2: Final synthesized answers with professional styling."""
        ws = self.wb.create_sheet("Synthesized Answers", 1)

        headers = ["#", "Question", "Synthesized Answer", "Sources", "Frags", "Notes"]
        col_widths = [5, 40, 80, 15, 8, 8]

        # Header row with styling
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(1, col, header)
            cell.font = self.header_font
            cell.fill = PatternFill("solid", fgColor=self.NAVY)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 25

        # Freeze header row
        ws.freeze_panes = 'A2'

        row = 2
        for qid, ca_data in self.accumulator.get('cumulative_answers', {}).items():
            # Alternating row colors
            row_fill = PatternFill("solid", fgColor=self.LIGHT_GRAY) if row % 2 == 0 else None

            # Row number
            ws.cell(row, 1, row - 1)
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 1).border = self.thin_border
            if row_fill:
                ws.cell(row, 1).fill = row_fill

            # Question
            ws.cell(row, 2, sanitize_for_excel(ca_data.get('question_text', '')))
            ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row, 2).border = self.thin_border
            if row_fill:
                ws.cell(row, 2).fill = row_fill

            # Use synthesized answer if available, else concatenate fragments
            synthesized = ca_data.get('synthesized_answer')
            if not synthesized:
                fragments = ca_data.get('fragments', [])
                if fragments:
                    synthesized = "\n\n---\n\n".join([sanitize_for_excel(f.get('text', '')) for f in fragments])
                else:
                    synthesized = "No answer found"
            else:
                synthesized = sanitize_for_excel(synthesized)

            # Answer
            ws.cell(row, 3, synthesized)
            ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row, 3).border = self.thin_border
            if row_fill:
                ws.cell(row, 3).fill = row_fill

            # Sources (pages)
            ws.cell(row, 4, ', '.join(map(str, ca_data.get('all_pages', []))))
            ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 4).font = Font(bold=True, color=self.NAVY)
            ws.cell(row, 4).border = self.thin_border
            if row_fill:
                ws.cell(row, 4).fill = row_fill

            # Fragment count
            frag_count = ca_data.get('fragment_count', 0)
            ws.cell(row, 5, frag_count)
            ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 5).border = self.thin_border
            if frag_count > 1:
                ws.cell(row, 5).font = Font(bold=True, color=self.ORANGE)
            if row_fill:
                ws.cell(row, 5).fill = row_fill

            # Footnote count
            ws.cell(row, 6, ca_data.get('footnote_count', 0))
            ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 6).border = self.thin_border
            if row_fill:
                ws.cell(row, 6).fill = row_fill

            # Dynamic row height based on answer length
            answer_len = len(synthesized) if synthesized else 0
            ws.row_dimensions[row].height = max(30, min(300, (answer_len // 80) * 15))

            row += 1

    def _create_fragments_sheet(self):
        """Sheet 3: All individual answer fragments with professional styling."""
        ws = self.wb.create_sheet("All Fragments", 2)

        headers = ["Fragment ID", "Question ID", "Window", "Pages", "Confidence", "Expert", "Fragment Text"]
        col_widths = [12, 15, 8, 15, 10, 25, 80]

        # Header row with styling
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(1, col, header)
            cell.font = self.header_font
            cell.fill = PatternFill("solid", fgColor=self.BLUE)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 25

        # Freeze header row
        ws.freeze_panes = 'A2'

        row = 2
        for qid, ca_data in self.accumulator.get('cumulative_answers', {}).items():
            for frag in ca_data.get('fragments', []):
                # Alternating row colors
                row_fill = PatternFill("solid", fgColor=self.LIGHT_GRAY) if row % 2 == 0 else None

                ws.cell(row, 1, frag.get('fragment_id', ''))
                ws.cell(row, 1).font = Font(size=9, color="666666")
                ws.cell(row, 1).border = self.thin_border
                if row_fill:
                    ws.cell(row, 1).fill = row_fill

                ws.cell(row, 2, qid)
                ws.cell(row, 2).border = self.thin_border
                if row_fill:
                    ws.cell(row, 2).fill = row_fill

                ws.cell(row, 3, frag.get('window_index', 0))
                ws.cell(row, 3).alignment = Alignment(horizontal='center')
                ws.cell(row, 3).border = self.thin_border
                if row_fill:
                    ws.cell(row, 3).fill = row_fill

                ws.cell(row, 4, ', '.join(map(str, frag.get('pages', []))))
                ws.cell(row, 4).font = Font(bold=True, color=self.NAVY)
                ws.cell(row, 4).alignment = Alignment(horizontal='center')
                ws.cell(row, 4).border = self.thin_border
                if row_fill:
                    ws.cell(row, 4).fill = row_fill

                confidence = frag.get('confidence', 0)
                ws.cell(row, 5, f"{confidence:.0%}")
                ws.cell(row, 5).alignment = Alignment(horizontal='center')
                ws.cell(row, 5).border = self.thin_border
                # Color code confidence
                if confidence >= 0.8:
                    ws.cell(row, 5).font = Font(bold=True, color=self.GREEN)
                elif confidence >= 0.5:
                    ws.cell(row, 5).font = Font(color=self.ORANGE)
                else:
                    ws.cell(row, 5).font = Font(color="EF4444")
                if row_fill:
                    ws.cell(row, 5).fill = row_fill

                ws.cell(row, 6, sanitize_for_excel(frag.get('expert_name', '')))
                ws.cell(row, 6).border = self.thin_border
                if row_fill:
                    ws.cell(row, 6).fill = row_fill

                ws.cell(row, 7, sanitize_for_excel(frag.get('text', '')))
                ws.cell(row, 7).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(row, 7).border = self.thin_border
                if row_fill:
                    ws.cell(row, 7).fill = row_fill

                row += 1

    def _create_footnotes_sheet(self):
        """Sheet 4: All individual footnotes with quotes and professional styling."""
        ws = self.wb.create_sheet("All Footnotes", 3)

        headers = ["Footnote ID", "Question ID", "Page", "Quote", "Window", "Fragment ID"]
        col_widths = [12, 15, 8, 80, 8, 12]

        # Header row with styling
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(1, col, header)
            cell.font = self.header_font
            cell.fill = PatternFill("solid", fgColor=self.GREEN)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 25

        # Freeze header row
        ws.freeze_panes = 'A2'

        row = 2
        for qid, ca_data in self.accumulator.get('cumulative_answers', {}).items():
            for fn in ca_data.get('footnotes', []):
                # Alternating row colors
                row_fill = PatternFill("solid", fgColor=self.LIGHT_GRAY) if row % 2 == 0 else None

                ws.cell(row, 1, fn.get('footnote_id', ''))
                ws.cell(row, 1).font = Font(size=9, color="666666")
                ws.cell(row, 1).border = self.thin_border
                if row_fill:
                    ws.cell(row, 1).fill = row_fill

                ws.cell(row, 2, qid)
                ws.cell(row, 2).border = self.thin_border
                if row_fill:
                    ws.cell(row, 2).fill = row_fill

                ws.cell(row, 3, fn.get('page', 0))
                ws.cell(row, 3).font = Font(bold=True, color=self.NAVY)
                ws.cell(row, 3).alignment = Alignment(horizontal='center')
                ws.cell(row, 3).border = self.thin_border
                if row_fill:
                    ws.cell(row, 3).fill = row_fill

                ws.cell(row, 4, sanitize_for_excel(fn.get('quote', '')))
                ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(row, 4).border = self.thin_border
                if row_fill:
                    ws.cell(row, 4).fill = row_fill

                ws.cell(row, 5, fn.get('window_index', 0))
                ws.cell(row, 5).alignment = Alignment(horizontal='center')
                ws.cell(row, 5).border = self.thin_border
                if row_fill:
                    ws.cell(row, 5).fill = row_fill

                ws.cell(row, 6, fn.get('fragment_id', ''))
                ws.cell(row, 6).font = Font(size=9, color="666666")
                ws.cell(row, 6).border = self.thin_border
                if row_fill:
                    ws.cell(row, 6).fill = row_fill

                row += 1

    def _create_sources_sheet(self):
        """Sheet 5: Study Guide with prioritized pages, visual density indicators, and raw page index."""
        from openpyxl.formatting.rule import DataBarRule

        ws = self.wb.create_sheet("Study Guide", 4)

        # Column widths for Study Guide layout
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25

        # Build comprehensive page data
        page_map = {}  # page -> list of question_ids
        page_fragments = {}  # page -> fragment count
        page_sections = {}  # page -> set of sections

        for qid, ca_data in self.accumulator.get('cumulative_answers', {}).items():
            # Extract section from question_id (format: SECTION_Q1)
            section = qid.split('_')[0] if '_' in qid else 'General'

            for page in ca_data.get('all_pages', []):
                if page not in page_map:
                    page_map[page] = []
                    page_fragments[page] = 0
                    page_sections[page] = set()
                if qid not in page_map[page]:
                    page_map[page].append(qid)
                page_fragments[page] += ca_data.get('fragment_count', 1)
                page_sections[page].add(section)

        # Calculate statistics
        total_pages = len(page_map)
        if total_pages == 0:
            # No data - show empty state
            ws.cell(1, 1, "No page references found in analysis")
            return

        ref_counts = [len(qids) for qids in page_map.values()]
        max_refs = max(ref_counts) if ref_counts else 0
        avg_refs = sum(ref_counts) / len(ref_counts) if ref_counts else 0

        # Categorize pages by priority
        high_priority = []  # >= 75th percentile
        medium_priority = []  # 25th-75th percentile
        low_priority = []  # < 25th percentile

        sorted_pages = sorted(page_map.keys(), key=lambda p: len(page_map[p]), reverse=True)
        threshold_high = max(1, int(len(sorted_pages) * 0.25))
        threshold_low = max(1, int(len(sorted_pages) * 0.75))

        for i, page in enumerate(sorted_pages):
            if i < threshold_high:
                high_priority.append(page)
            elif i < threshold_low:
                medium_priority.append(page)
            else:
                low_priority.append(page)

        # ============================================================
        # SECTION 1: STUDY GUIDE HEADER
        # ============================================================
        row = 1
        ws.cell(row, 1, "📚 STUDY GUIDE")
        ws.cell(row, 1).font = Font(size=20, bold=True, color=self.NAVY)
        ws.merge_cells('A1:F1')
        ws.row_dimensions[row].height = 35

        row = 2
        ws.cell(row, 1, "Prioritized Page Reference Guide for Focused Study")
        ws.cell(row, 1).font = Font(size=12, italic=True, color="666666")
        ws.merge_cells('A2:F2')

        # ============================================================
        # SECTION 2: SUMMARY STATISTICS
        # ============================================================
        row = 4
        ws.cell(row, 1, "OVERVIEW")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.NAVY)
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=self.NAVY)
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=self.NAVY)
        ws.merge_cells(f'C{row}:D{row}')

        stats_data = [
            ("Total Pages Referenced", total_pages, "Max Refs on Single Page", max_refs),
            ("High Priority Pages", len(high_priority), "Avg Questions/Page", f"{avg_refs:.1f}"),
            ("Medium Priority Pages", len(medium_priority), "Total Fragments", sum(page_fragments.values())),
        ]

        for label1, val1, label2, val2 in stats_data:
            row += 1
            ws.cell(row, 1, label1)
            ws.cell(row, 1).font = self.label_font
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.LIGHT_GRAY)
            ws.cell(row, 1).border = self.thin_border
            ws.cell(row, 2, val1)
            ws.cell(row, 2).font = Font(bold=True, color=self.NAVY)
            ws.cell(row, 2).alignment = Alignment(horizontal='center')
            ws.cell(row, 2).border = self.thin_border
            ws.cell(row, 3, label2)
            ws.cell(row, 3).font = self.label_font
            ws.cell(row, 3).fill = PatternFill("solid", fgColor=self.LIGHT_GRAY)
            ws.cell(row, 3).border = self.thin_border
            ws.cell(row, 4, val2)
            ws.cell(row, 4).font = Font(bold=True, color=self.NAVY)
            ws.cell(row, 4).alignment = Alignment(horizontal='center')
            ws.cell(row, 4).border = self.thin_border

        # ============================================================
        # SECTION 3: HIGH PRIORITY PAGES (STUDY FIRST!)
        # ============================================================
        row += 2
        ws.cell(row, 1, "🔴 HIGH PRIORITY - Study These First!")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="DC2626")  # Red
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 25

        row += 1
        priority_headers = ["Page", "Questions", "Sections", "Density", "Priority"]
        for col, header in enumerate(priority_headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color=self.NAVY, size=10)
            cell.fill = PatternFill("solid", fgColor="FECACA")  # Light red
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border

        density_start_row = row + 1
        for page in sorted(high_priority, key=lambda p: len(page_map[p]), reverse=True)[:10]:  # Top 10
            row += 1
            ref_count = len(page_map[page])
            sections = ', '.join(sorted(page_sections.get(page, set())))

            ws.cell(row, 1, f"Page {page}")
            ws.cell(row, 1).font = Font(bold=True, color=self.NAVY, size=11)
            ws.cell(row, 1).alignment = Alignment(horizontal='center')
            ws.cell(row, 1).border = self.thin_border

            ws.cell(row, 2, ref_count)
            ws.cell(row, 2).alignment = Alignment(horizontal='center')
            ws.cell(row, 2).border = self.thin_border

            ws.cell(row, 3, sections[:40] + ('...' if len(sections) > 40 else ''))
            ws.cell(row, 3).border = self.thin_border

            # Density score (for data bar)
            density = ref_count
            ws.cell(row, 4, density)
            ws.cell(row, 4).alignment = Alignment(horizontal='center')
            ws.cell(row, 4).border = self.thin_border

            ws.cell(row, 5, "⭐⭐⭐")
            ws.cell(row, 5).alignment = Alignment(horizontal='center')
            ws.cell(row, 5).font = Font(color="DC2626")
            ws.cell(row, 5).border = self.thin_border

        # Add data bar to density column for high priority
        if row > density_start_row:
            data_bar_rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='num', end_value=max_refs,
                color="DC2626"  # Red
            )
            ws.conditional_formatting.add(f'D{density_start_row}:D{row}', data_bar_rule)

        # ============================================================
        # SECTION 4: MEDIUM PRIORITY PAGES
        # ============================================================
        row += 2
        ws.cell(row, 1, "🟡 MEDIUM PRIORITY - Review After High Priority")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.ORANGE)
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 25

        row += 1
        for col, header in enumerate(priority_headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color=self.ORANGE, size=10)
            cell.fill = PatternFill("solid", fgColor=self.LIGHT_ORANGE)
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border

        density_start_row = row + 1
        for page in sorted(medium_priority, key=lambda p: len(page_map[p]), reverse=True)[:15]:  # Top 15
            row += 1
            ref_count = len(page_map[page])
            sections = ', '.join(sorted(page_sections.get(page, set())))

            ws.cell(row, 1, f"Page {page}")
            ws.cell(row, 1).font = Font(bold=True, color=self.NAVY)
            ws.cell(row, 1).alignment = Alignment(horizontal='center')
            ws.cell(row, 1).border = self.thin_border

            ws.cell(row, 2, ref_count)
            ws.cell(row, 2).alignment = Alignment(horizontal='center')
            ws.cell(row, 2).border = self.thin_border

            ws.cell(row, 3, sections[:40] + ('...' if len(sections) > 40 else ''))
            ws.cell(row, 3).border = self.thin_border

            density = ref_count
            ws.cell(row, 4, density)
            ws.cell(row, 4).alignment = Alignment(horizontal='center')
            ws.cell(row, 4).border = self.thin_border

            ws.cell(row, 5, "⭐⭐")
            ws.cell(row, 5).alignment = Alignment(horizontal='center')
            ws.cell(row, 5).font = Font(color=self.ORANGE)
            ws.cell(row, 5).border = self.thin_border

        # Add data bar for medium priority
        if row > density_start_row:
            data_bar_rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='num', end_value=max_refs,
                color=self.ORANGE
            )
            ws.conditional_formatting.add(f'D{density_start_row}:D{row}', data_bar_rule)

        # ============================================================
        # SECTION 5: LOW PRIORITY (Quick Reference)
        # ============================================================
        row += 2
        ws.cell(row, 1, "🟢 LOW PRIORITY - Optional/Supplementary")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.GREEN)
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 25

        row += 1
        # Compact format for low priority
        ws.cell(row, 1, "Pages")
        ws.cell(row, 1).font = Font(bold=True, color=self.GREEN)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="DCFCE7")  # Light green
        ws.cell(row, 1).border = self.thin_border

        low_pages_str = ', '.join([f"{p} ({len(page_map[p])})" for p in sorted(low_priority)])
        ws.cell(row, 2, low_pages_str if low_pages_str else "None")
        ws.cell(row, 2).alignment = Alignment(wrap_text=True)
        ws.cell(row, 2).border = self.thin_border
        ws.merge_cells(f'B{row}:F{row}')

        # ============================================================
        # SECTION 6: RAW PAGE INDEX (Preserved Data)
        # ============================================================
        row += 3
        ws.cell(row, 1, "📋 COMPLETE PAGE INDEX (Raw Data)")
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.NAVY)
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 25

        row += 1
        index_headers = ["Page", "Question Count", "Questions Referencing This Page", "Sections", "Fragment Count"]
        for col, header in enumerate(index_headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = PatternFill("solid", fgColor=self.BLUE)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        ws.row_dimensions[row].height = 22

        for page in sorted(page_map.keys()):
            row += 1
            row_fill = PatternFill("solid", fgColor=self.LIGHT_GRAY) if row % 2 == 0 else None

            ws.cell(row, 1, page)
            ws.cell(row, 1).font = Font(bold=True, color=self.NAVY, size=11)
            ws.cell(row, 1).alignment = Alignment(horizontal='center')
            ws.cell(row, 1).border = self.thin_border
            if row_fill:
                ws.cell(row, 1).fill = row_fill

            ref_count = len(page_map[page])
            ws.cell(row, 2, ref_count)
            ws.cell(row, 2).alignment = Alignment(horizontal='center')
            ws.cell(row, 2).border = self.thin_border
            if ref_count >= 5:
                ws.cell(row, 2).font = Font(bold=True, color=self.ORANGE)
            if row_fill:
                ws.cell(row, 2).fill = row_fill

            ws.cell(row, 3, ', '.join(page_map[page]))
            ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row, 3).border = self.thin_border
            if row_fill:
                ws.cell(row, 3).fill = row_fill

            sections = ', '.join(sorted(page_sections.get(page, set())))
            ws.cell(row, 4, sections)
            ws.cell(row, 4).border = self.thin_border
            if row_fill:
                ws.cell(row, 4).fill = row_fill

            ws.cell(row, 5, page_fragments.get(page, 0))
            ws.cell(row, 5).alignment = Alignment(horizontal='center')
            ws.cell(row, 5).border = self.thin_border
            if row_fill:
                ws.cell(row, 5).fill = row_fill
