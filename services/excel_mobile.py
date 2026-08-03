"""
Mobile-first sizing for every BidBrief Excel export.

The primary reader is the iOS app's QuickLook preview on a phone: wide desktop
columns force endless horizontal panning there. Desktop Excel users can resize
freely, so phone-friendly defaults cost nothing on desktop.

Applied by all generators (analysis dashboard, BestPrep, Smart Analysis,
CityScraper) right before the workbook is saved.
"""

import math

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, range_boundaries

# Max column width (openpyxl width units ≈ characters). ~42 keeps two columns
# readable side-by-side on a phone in portrait without desktop feeling cramped.
MAX_COL_WIDTH = 42
DEFAULT_COL_WIDTH = 24
# Ceiling for a column that holds prose too long to fit the clamp without
# hitting Excel's row-height limit. Wider than phone-ideal, but visible beats
# tidy — a clipped answer is a wrong answer.
LONG_FORM_COL_WIDTH = 72

# Excel's own ceiling for a row is 409.5 points.
MAX_ROW_HEIGHT = 409
# Calibri 11 renders one wrapped line in roughly this many points.
POINTS_PER_LINE = 14.5
# Width Excel uses for a column nobody sized.
IMPLICIT_COL_WIDTH = 8.43


def _effective_width(ws, cell, merged_spans):
    """Characters that fit on one line in this cell.

    A merged cell spans several columns, so its usable width is their sum —
    measuring only the anchor column would badly overestimate the wrapping and
    produce absurdly tall rows on the summary sheets.
    """
    col = cell.column
    span = merged_spans.get((cell.row, col))
    columns = range(span[0], span[1] + 1) if span else (col,)
    total = 0.0
    for index in columns:
        dim = ws.column_dimensions.get(get_column_letter(index))
        total += (dim.width if dim and dim.width else IMPLICIT_COL_WIDTH)
    return max(total, 4.0)


def _lines_needed(text, width):
    """Wrapped line count for a cell's text, honouring explicit newlines."""
    lines = 0
    for paragraph in str(text).split('\n'):
        lines += max(1, math.ceil(len(paragraph) / max(width - 1, 1)))
    return lines


def autosize_rows(ws):
    """Grow every row so wrapped text is fully visible.

    MUST run AFTER the column clamp: measuring against the pre-clamp width
    under-sizes the row, which is precisely how a long answer ended up wrapped
    into a 42-character column but cut off by a hard-coded 55pt height.

    Only ever grows a row — deliberate header/banner heights are preserved.
    """
    merged_spans = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        for row in range(min_row, max_row + 1):
            merged_spans[(row, min_col)] = (min_col, max_col)

    for row_cells in ws.iter_rows():
        if not row_cells:
            continue
        row_index = row_cells[0].row
        needed = 0
        for cell in row_cells:
            if cell.value is None:
                continue
            alignment = cell.alignment
            if not (alignment and alignment.wrap_text):
                continue
            width = _effective_width(ws, cell, merged_spans)
            needed = max(needed, _lines_needed(cell.value, width))

        if needed <= 1:
            continue

        target = min(needed * POINTS_PER_LINE, MAX_ROW_HEIGHT)
        dim = ws.row_dimensions[row_index]
        if dim.height is None or dim.height < target:
            dim.height = round(target, 1)


def _relieve_overflowing_columns(ws, clamped_cols):
    """Widen a clamped column when 42 characters would clip its content.

    A row cannot exceed 409.5 points in Excel. Narrowing a column makes text
    wrap into more lines, so an aggressive clamp can push a long answer past
    that ceiling and silently hide the tail. Phone-friendly sizing is never
    worth losing data, so any column holding prose too long for the clamp is
    widened just enough to fit — and no further.
    """
    max_lines = MAX_ROW_HEIGHT / POINTS_PER_LINE
    for letter in list(clamped_cols):
        longest = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            for paragraph in str(cell.value).split('\n'):
                longest = max(longest, len(paragraph))
        if not longest:
            continue
        required = (longest / max_lines) + 1
        if required > MAX_COL_WIDTH:
            ws.column_dimensions[letter].width = min(
                math.ceil(required), LONG_FORM_COL_WIDTH)
    return ws


def mobile_optimize(wb):
    """Clamp column widths, wrap clamped columns, and set fit-to-width printing."""
    for ws in wb.worksheets:
        clamped_cols = set()
        for letter, dim in list(ws.column_dimensions.items()):
            if dim.width and dim.width > MAX_COL_WIDTH:
                dim.width = MAX_COL_WIDTH
                clamped_cols.add(letter)

        _relieve_overflowing_columns(ws, clamped_cols)

        # Text that used to fit a wide column must wrap in the clamped one.
        for letter in clamped_cols:
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                current = cell.alignment or Alignment()
                if not current.wrap_text:
                    cell.alignment = Alignment(
                        horizontal=current.horizontal,
                        vertical=current.vertical or 'top',
                        wrap_text=True,
                        indent=current.indent or 0,
                    )

        # Rows last: the height depends on the FINAL column widths, so this has
        # to see the clamped values, not the ones the generator asked for.
        autosize_rows(ws)

        # Phone-friendly view + print defaults.
        ws.sheet_view.zoomScale = 100
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb
