"""
Smart Analysis Excel Generator — 4-sheet export.

Sheet 1: Executive Summary  (metadata, key insights, professional assessments)
Sheet 2: Findings           (risks, opportunities, ambiguities, contradictions)
Sheet 3: Recommendations    (follow-up questions, strategic recommendations)
Sheet 4: User Questions     (user question / response pairs, if any)

Style matches services/excel_dashboard.py aesthetic.
"""

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import SmartAnalysisResult

logger = logging.getLogger(__name__)

# ── Colours (match existing excel_dashboard.py palette) ──────────────────────
_NAVY = '1E3A8A'
_BLUE = '3B82F6'
_GREEN = '22C55E'
_AMBER = 'F59E0B'
_RED = 'EF4444'
_DARK_RED = 'DC2626'
_LIGHT_BLUE_BG = 'EFF6FF'
_LIGHT_GREY_BG = 'F8FAFC'
_ALT_ROW = 'F1F5F9'
_WHITE = 'FFFFFF'

# ── Reusable style factories ──────────────────────────────────────────────────
def _font(bold=False, size=11, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _border():
    thin = Side(style='thin', color='D1D5DB')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(wrap=True, h='left', v='top'):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v, indent=1)


_HEADER_FONT = _font(bold=True, size=13, color=_WHITE)
_HEADER_FILL = _fill(_NAVY)
_SECTION_FONT = _font(bold=True, size=11, color=_WHITE)
_SECTION_FILL = _fill(_BLUE)
_LABEL_FONT = _font(bold=True, size=10)
_DATA_FONT = _font(size=10)
_BORDER = _border()


# ── Severity colour map ───────────────────────────────────────────────────────
_SEVERITY_FILLS = {
    'critical': _fill(_DARK_RED),
    'high':     _fill(_RED),
    'medium':   _fill(_AMBER),
    'low':      _fill(_GREEN),
    'info':     _fill(_BLUE),
}
_SEVERITY_FONTS = {
    'critical': _font(bold=True, color=_WHITE, size=10),
    'high':     _font(bold=True, color=_WHITE, size=10),
    'medium':   _font(bold=True, color='1F2937', size=10),
    'low':      _font(bold=True, color=_WHITE, size=10),
    'info':     _font(bold=True, color=_WHITE, size=10),
}


class SmartAnalysisExcelGenerator:
    def __init__(self, result: SmartAnalysisResult):
        self.result = result
        self.wb = Workbook()

    def generate(self) -> io.BytesIO:
        self.wb.remove(self.wb.active)          # Remove default sheet

        self._sheet_summary()
        self._sheet_findings()
        self._sheet_recommendations()
        if self.result.user_question_responses:
            self._sheet_user_questions()

        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Sheet 1: Executive Summary ────────────────────────────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet('Executive Summary')
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 80
        row = 1

        # Title header
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'BidBrief Smart Analysis — Executive Summary'
        ws[f'A{row}'].font = _HEADER_FONT
        ws[f'A{row}'].fill = _HEADER_FILL
        ws[f'A{row}'].alignment = _align(h='center', wrap=False)
        ws.row_dimensions[row].height = 36
        row += 1

        # Metadata
        meta = [
            ('Document', self.result.document_name),
            ('Document Type', self.result.document_type_label or self.result.document_type),
            ('Analysis Coverage', self.result.analysis_completeness.title()),
            ('Generated', _fmt_date(self.result.generated_at)),
        ]
        for label, value in meta:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = _LABEL_FONT
            ws[f'A{row}'].fill = _fill(_LIGHT_GREY_BG)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = _DATA_FONT
            for col in 'AB':
                ws[f'{col}{row}'].border = _BORDER
                ws[f'{col}{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1

        # Professional Assessments
        if self.result.assessments:
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = 'Professional Assessments'
            ws[f'A{row}'].font = _SECTION_FONT
            ws[f'A{row}'].fill = _SECTION_FILL
            ws[f'A{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 24
            row += 1

            for a in self.result.assessments:
                ws[f'A{row}'] = a.category
                ws[f'A{row}'].font = _LABEL_FONT
                ws[f'A{row}'].fill = _fill(_LIGHT_BLUE_BG)
                rating_cell = ws[f'B{row}']
                rating_cell.value = f"{a.rating}  —  {a.rationale}  (confidence: {a.confidence})"
                rating_cell.font = _DATA_FONT
                for col in 'AB':
                    ws[f'{col}{row}'].border = _BORDER
                    ws[f'{col}{row}'].alignment = _align()
                ws.row_dimensions[row].height = 30
                row += 1

            row += 1

        # Key Insights
        if self.result.key_insights:
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = 'Key Insights'
            ws[f'A{row}'].font = _SECTION_FONT
            ws[f'A{row}'].fill = _SECTION_FILL
            ws[f'A{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 24
            row += 1

            for i, insight in enumerate(self.result.key_insights, 1):
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = f'{i}.  {insight}'
                ws[f'A{row}'].font = _DATA_FONT
                ws[f'A{row}'].fill = _fill(_ALT_ROW if i % 2 == 0 else _WHITE)
                ws[f'A{row}'].border = _BORDER
                ws[f'A{row}'].alignment = _align()
                ws.row_dimensions[row].height = 40
                row += 1

            row += 1

        # Executive Summary prose
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'Executive Summary'
        ws[f'A{row}'].font = _SECTION_FONT
        ws[f'A{row}'].fill = _SECTION_FILL
        ws[f'A{row}'].alignment = _align(wrap=False)
        ws.row_dimensions[row].height = 24
        row += 1

        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = self.result.executive_summary
        ws[f'A{row}'].font = _DATA_FONT
        ws[f'A{row}'].border = _BORDER
        ws[f'A{row}'].alignment = _align()
        ws.row_dimensions[row].height = 200
        row += 1

    # ── Sheet 2: Findings ─────────────────────────────────────────────────────
    def _sheet_findings(self):
        ws = self.wb.create_sheet('Findings')
        ws.column_dimensions['A'].width = 6    # Severity badge
        ws.column_dimensions['B'].width = 28   # Title
        ws.column_dimensions['C'].width = 70   # Description
        ws.column_dimensions['D'].width = 40   # Evidence
        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'BidBrief Smart Analysis — Findings'
        ws[f'A{row}'].font = _HEADER_FONT
        ws[f'A{row}'].fill = _HEADER_FILL
        ws[f'A{row}'].alignment = _align(h='center', wrap=False)
        ws.row_dimensions[row].height = 36
        row += 1

        categories = [
            ('Risks',           self.result.risks),
            ('Opportunities',   self.result.opportunities),
            ('Ambiguities',     self.result.ambiguities),
            ('Contradictions',  self.result.contradictions),
        ]

        for cat_name, items in categories:
            if not items:
                continue

            row += 1
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = cat_name
            ws[f'A{row}'].font = _SECTION_FONT
            ws[f'A{row}'].fill = _SECTION_FILL
            ws[f'A{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 24
            row += 1

            # Column headers
            for col, label in zip('ABCD', ['Sev.', 'Title', 'Description', 'Evidence']):
                ws[f'{col}{row}'] = label
                ws[f'{col}{row}'].font = _LABEL_FONT
                ws[f'{col}{row}'].fill = _fill(_LIGHT_GREY_BG)
                ws[f'{col}{row}'].border = _BORDER
                ws[f'{col}{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 20
            row += 1

            for item in items:
                sev = item.severity.lower()
                sev_fill = _SEVERITY_FILLS.get(sev, _fill(_AMBER))
                sev_font = _SEVERITY_FONTS.get(sev, _LABEL_FONT)

                ws[f'A{row}'] = sev.upper()[:4]
                ws[f'A{row}'].font = sev_font
                ws[f'A{row}'].fill = sev_fill
                ws[f'A{row}'].border = _BORDER
                ws[f'A{row}'].alignment = _align(h='center', wrap=False)

                ws[f'B{row}'] = item.title
                ws[f'B{row}'].font = _LABEL_FONT
                ws[f'B{row}'].border = _BORDER
                ws[f'B{row}'].alignment = _align()

                ws[f'C{row}'] = item.description
                ws[f'C{row}'].font = _DATA_FONT
                ws[f'C{row}'].border = _BORDER
                ws[f'C{row}'].alignment = _align()

                evidence_text = '\n'.join(f'• {e}' for e in item.evidence) if item.evidence else ''
                if item.page_refs:
                    evidence_text += f'\nPages: {item.page_refs}'
                ws[f'D{row}'] = evidence_text
                ws[f'D{row}'].font = _DATA_FONT
                ws[f'D{row}'].border = _BORDER
                ws[f'D{row}'].alignment = _align()

                ws.row_dimensions[row].height = 60
                row += 1

    # ── Sheet 3: Recommendations ──────────────────────────────────────────────
    def _sheet_recommendations(self):
        ws = self.wb.create_sheet('Recommendations')
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 90
        row = 1

        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'BidBrief Smart Analysis — Recommendations'
        ws[f'A{row}'].font = _HEADER_FONT
        ws[f'A{row}'].fill = _HEADER_FILL
        ws[f'A{row}'].alignment = _align(h='center', wrap=False)
        ws.row_dimensions[row].height = 36
        row += 1

        if self.result.strategic_recommendations:
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = 'Strategic Recommendations'
            ws[f'A{row}'].font = _SECTION_FONT
            ws[f'A{row}'].fill = _SECTION_FILL
            ws[f'A{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 24
            row += 1

            for i, rec in enumerate(self.result.strategic_recommendations, 1):
                ws[f'A{row}'] = str(i)
                ws[f'A{row}'].font = _LABEL_FONT
                ws[f'A{row}'].fill = _fill(_LIGHT_BLUE_BG)
                ws[f'A{row}'].border = _BORDER
                ws[f'A{row}'].alignment = _align(h='center', wrap=False)

                ws[f'B{row}'] = rec
                ws[f'B{row}'].font = _DATA_FONT
                ws[f'B{row}'].border = _BORDER
                ws[f'B{row}'].alignment = _align()
                ws.row_dimensions[row].height = 50
                row += 1

        if self.result.follow_up_questions:
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = 'Follow-Up Questions for Further Investigation'
            ws[f'A{row}'].font = _SECTION_FONT
            ws[f'A{row}'].fill = _SECTION_FILL
            ws[f'A{row}'].alignment = _align(wrap=False)
            ws.row_dimensions[row].height = 24
            row += 1

            for i, q in enumerate(self.result.follow_up_questions, 1):
                ws[f'A{row}'] = str(i)
                ws[f'A{row}'].font = _LABEL_FONT
                ws[f'A{row}'].fill = _fill(_LIGHT_GREY_BG)
                ws[f'A{row}'].border = _BORDER
                ws[f'A{row}'].alignment = _align(h='center', wrap=False)

                ws[f'B{row}'] = q
                ws[f'B{row}'].font = _DATA_FONT
                ws[f'B{row}'].border = _BORDER
                ws[f'B{row}'].alignment = _align()
                ws.row_dimensions[row].height = 45
                row += 1

    # ── Sheet 4: User Questions ───────────────────────────────────────────────
    def _sheet_user_questions(self):
        ws = self.wb.create_sheet('Your Questions')
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 12
        row = 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = 'BidBrief Smart Analysis — Your Questions'
        ws[f'A{row}'].font = _HEADER_FONT
        ws[f'A{row}'].fill = _HEADER_FILL
        ws[f'A{row}'].alignment = _align(h='center', wrap=False)
        ws.row_dimensions[row].height = 36
        row += 1

        # Column headers
        row += 1
        for col, label in zip('ABC', ['Question', 'Response', 'Confidence']):
            ws[f'{col}{row}'] = label
            ws[f'{col}{row}'].font = _LABEL_FONT
            ws[f'{col}{row}'].fill = _fill(_LIGHT_GREY_BG)
            ws[f'{col}{row}'].border = _BORDER
            ws[f'{col}{row}'].alignment = _align(wrap=False)
        ws.row_dimensions[row].height = 20
        row += 1

        for i, resp in enumerate(self.result.user_question_responses):
            fill = _fill(_ALT_ROW if i % 2 else _WHITE)

            ws[f'A{row}'] = resp.get('question', '')
            ws[f'B{row}'] = resp.get('response', '')
            ws[f'C{row}'] = resp.get('confidence', '').title()

            for col in 'ABC':
                ws[f'{col}{row}'].font = _DATA_FONT
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = _BORDER
                ws[f'{col}{row}'].alignment = _align()
            ws.row_dimensions[row].height = 70
            row += 1


def _fmt_date(iso_str: str) -> str:
    try:
        dt = datetime.fromisoformat(iso_str.replace('Z', '+00:00'))
        return dt.strftime('%B %d, %Y at %I:%M %p UTC')
    except Exception:
        return iso_str
