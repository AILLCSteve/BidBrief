"""
Executive Excel Report Package - 4-Sheet Professional Format
Industry-grade formatting with openpyxl for BidBrief Analysis Results

Sheet 1: Executive Summary - Analysis statistics and key project requirements
Sheet 2: Detailed Results - All questions with answers in unified table
Sheet 3: By Section - Questions grouped by section with clear headers
Sheet 4: Footnotes - All footnotes and citation references
"""
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_for_excel(text):
    """
    Remove or replace characters that are illegal in Excel worksheet cells.
    Excel doesn't allow control characters (0x00-0x1F except tab, newline, carriage return).
    """
    if not text:
        return text
    if not isinstance(text, str):
        return str(text)

    # Remove illegal XML characters (control chars except tab, newline, carriage return)
    # These cause "cannot be used in worksheets" errors
    illegal_pattern = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
    cleaned = illegal_pattern.sub('', text)

    # Also replace other problematic characters
    # \u001f is Unit Separator - replace with space
    cleaned = cleaned.replace('\u001f', ' ')

    return cleaned


class ExcelDashboardGenerator:
    """Generate executive-ready Excel report package with 4 professional sheets"""

    # Professional color scheme - BidBrief Navy/Blue branding
    NAVY = "104090"
    BLUE = "5E86D0"
    LIGHT_BLUE = "E8EEF7"
    GREEN = "22C55E"
    GREEN_LIGHT = "DCFCE7"
    RED = "EF4444"
    RED_LIGHT = "FEE2E2"
    ORANGE = "F59E0B"
    ORANGE_LIGHT = "FEF3C7"
    GRAY = "6B7280"
    GRAY_LIGHT = "F3F4F6"
    WHITE = "FFFFFF"

    # Fills
    HEADER_FILL = PatternFill(start_color="104090", end_color="104090", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="5E86D0", end_color="5E86D0", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    ANSWERED_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    UNANSWERED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    STAT_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    TITLE_FILL = PatternFill(start_color="104090", end_color="104090", fill_type="solid")

    # Fonts
    TITLE_FONT = Font(name='Calibri', size=24, bold=True, color="FFFFFF")
    HEADER_FONT = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    SUBHEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    SECTION_FONT = Font(name='Calibri', size=12, bold=True, color="104090")
    DATA_FONT = Font(name='Calibri', size=11, color="374151")
    DATA_FONT_BOLD = Font(name='Calibri', size=11, bold=True, color="1F2937")
    STAT_LABEL_FONT = Font(name='Calibri', size=11, color="374151")
    STAT_VALUE_FONT = Font(name='Calibri', size=11, bold=True, color="104090")

    # Borders
    BORDER_THIN = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Key requirement mappings - questions to look for
    KEY_REQUIREMENT_PATTERNS = {
        'Project Name': ['project name', 'project title', 'contract name', 'project no', 'project number'],
        'Owner/Agency': ['owner', 'agency', 'municipality', 'city of', 'county of', 'district'],
        'Engineer': ['engineer', 'designer', 'architect', 'design firm', 'engineering firm'],
        'Timeline': ['timeline', 'project duration', 'start date', 'completion date', 'schedule'],
        'Scope': ['scope', 'linear feet', 'pipe diameter', 'total footage', 'project scope'],
        'Bid Deadline': ['bid deadline', 'bid due', 'submission deadline', 'bid opening'],
        'Payment': ['payment', 'progress payment', 'retention', 'payment terms'],
        'Warranty': ['warranty', 'guarantee period', 'warranty period'],
        'Liquidated Damages': ['liquidated damages', 'delay damages', 'penalty'],
        'Bonding': ['bond', 'bid bond', 'performance bond', 'payment bond'],
        'Certifications': ['certification', 'nassco', 'pacp', 'required certifications', 'operator certification'],
        'Insurance': ['insurance', 'liability', 'coverage requirements'],
        'Location': ['location', 'project location', 'city', 'municipality', 'address'],
    }

    def __init__(self, analysis_result, is_partial=False, api_key_requirements=None, optimized_scan_data=None, unanswered_pass_data=None, rag_data=None):
        self.result = analysis_result
        self.is_partial = is_partial
        self.optimized_scan_data = optimized_scan_data  # V2 pipeline document navigator data
        self.unanswered_pass_data = unanswered_pass_data  # V2 pipeline unanswered questions pass data
        self.rag_data = rag_data  # V2 pipeline RAG analysis data
        self.wb = Workbook()
        self.footnotes = self._collect_footnotes()
        # Use API-extracted key requirements if provided, otherwise extract from sections
        if api_key_requirements and len(api_key_requirements) > 0:
            self.key_requirements = self._format_api_key_requirements(api_key_requirements)
        else:
            self.key_requirements = self._extract_key_requirements()

    def generate(self):
        """Generate complete report package with V2 Pipeline sheets"""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Detect if V2 Pipeline (HOTDOG7ATE) was used
        is_v2_pipeline = self.optimized_scan_data is not None

        self._create_executive_summary()      # Sheet 1: Executive Summary

        # Sheet 2: Optimized Scan (V2 Pipeline - after Summary)
        if is_v2_pipeline:
            self._create_optimized_scan_sheet()

        self._create_detailed_results()       # Sheet 3: Detailed Results
        self._create_by_section()             # Sheet 4: By Section

        # Dynamic Intelligence sheet — document-specific tables sensed from THIS
        # document's results (shape varies per document by design)
        if self.result.get('dynamic_tables'):
            self._create_dynamic_intelligence_sheet()

        # Visual Intelligence sheet — findings from the opt-in Layer 0.5 vision
        # pass over drawing/map/photo-heavy pages (absent unless it ran)
        if self.result.get('visual_findings'):
            self._create_visual_intelligence_sheet()

        # Sheet 5: Unanswered Questions Pass (V2 Pipeline)
        if is_v2_pipeline:
            self._create_unanswered_pass_sheet()

        self._create_footnotes_sheet()        # Sheet 6: Footnotes

        # Sheet 7: Deep RAG (V2 Pipeline - at end)
        if is_v2_pipeline:
            self._create_rag_sheet()

        from services.excel_mobile import mobile_optimize
        mobile_optimize(self.wb)

        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _create_dynamic_intelligence_sheet(self):
        """Render the dynamic intelligence tables with real columns per table."""
        from openpyxl.utils import get_column_letter

        ws = self.wb.create_sheet('Document Intelligence')
        tables = self.result.get('dynamic_tables') or []
        max_cols = max(max((len(t.get('columns') or []) for t in tables), default=1), 2)
        last_col = get_column_letter(max_cols)
        for c in range(1, max_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 28

        row = 1
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws[f'A{row}']
        cell.value = 'Document Intelligence — Dynamic Tables'
        cell.font = self.HEADER_FONT
        cell.fill = self.TITLE_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 32
        row += 1

        focus = self.result.get('intelligence_focus') or ''
        if focus:
            ws.merge_cells(f'A{row}:{last_col}{row}')
            cell = ws[f'A{row}']
            cell.value = sanitize_for_excel(focus)
            cell.font = Font(name='Calibri', size=10, italic=True, color='374151')
            cell.fill = self.STAT_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 44
            row += 2

        for t in tables:
            columns = t.get('columns') or []
            if not columns:
                continue
            end_col = get_column_letter(len(columns))

            ws.merge_cells(f'A{row}:{end_col}{row}')
            cell = ws[f'A{row}']
            cell.value = sanitize_for_excel(t.get('title', 'Table'))
            cell.font = self.SECTION_FONT
            cell.fill = self.SECTION_FILL
            ws.row_dimensions[row].height = 22
            row += 1

            why = t.get('why_relevant') or ''
            if why:
                ws.merge_cells(f'A{row}:{end_col}{row}')
                cell = ws[f'A{row}']
                cell.value = sanitize_for_excel(why)
                cell.font = Font(name='Calibri', size=9, italic=True, color='6B7280')
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1

            for idx, col in enumerate(columns, 1):
                cell = ws[f'{get_column_letter(idx)}{row}']
                cell.value = sanitize_for_excel(col.get('label', col.get('key', '')))
                cell.font = self.SUBHEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER_THIN
                cell.alignment = Alignment(vertical='center')
            row += 1

            for r_i, data_row in enumerate(t.get('rows') or []):
                for idx, col in enumerate(columns, 1):
                    cell = ws[f'{get_column_letter(idx)}{row}']
                    cell.value = sanitize_for_excel(str(data_row.get(col.get('key', ''), '')))
                    cell.font = self.DATA_FONT
                    if r_i % 2:
                        cell.fill = self.ALT_ROW_FILL
                    cell.border = self.BORDER_THIN
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = 26
                row += 1

            for insight in t.get('insights') or []:
                ws.merge_cells(f'A{row}:{end_col}{row}')
                cell = ws[f'A{row}']
                cell.value = sanitize_for_excel(f'💡 {insight}')
                cell.font = Font(name='Calibri', size=9, italic=True, color='104090')
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1

            row += 1

    def _create_visual_intelligence_sheet(self):
        """Visual Intelligence — one row per drawing/map/photo finding from the
        opt-in Layer 0.5 vision pass. Mirrors the web results sheet."""
        ws = self.wb.create_sheet('Visual Intelligence')
        findings = self.result.get('visual_findings') or []

        headers = ['Page', 'Type', 'Title', 'What It Shows',
                   'Labels, Dimensions & Callouts', 'Key Facts']
        widths = [7, 12, 28, 55, 50, 50]

        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = 'VISUAL INTELLIGENCE — Drawings, Maps & Imagery'
        title_cell.font = self.HEADER_FONT
        title_cell.fill = self.TITLE_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:F2')
        note = ws['A2']
        note.value = (f'{len(findings)} visual-heavy page(s) deep-processed with AI vision, '
                      'in addition to (never instead of) the standard text analysis.')
        note.font = Font(name='Calibri', size=10, italic=True, color='374151')
        note.fill = self.STAT_FILL
        note.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[2].height = 24

        row = 3
        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.SUBHEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER_THIN
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[row].height = 24
        row += 1

        for idx, f in enumerate(findings):
            is_alt = idx % 2 == 1
            key_facts = f.get('key_facts') or []
            values = [
                f.get('page', ''),
                str(f.get('kind', '') or '').title(),
                sanitize_for_excel(f.get('title') or '-'),
                sanitize_for_excel(f.get('description') or '-'),
                sanitize_for_excel(f.get('extracted_text') or '-'),
                sanitize_for_excel('\n'.join(f'• {fact}' for fact in key_facts) or '-'),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col, value)
                cell.font = self.DATA_FONT_BOLD if col == 1 else self.DATA_FONT
                cell.border = self.BORDER_THIN
                if is_alt:
                    cell.fill = self.ALT_ROW_FILL
                cell.alignment = Alignment(
                    horizontal='center' if col in (1, 2) else 'left',
                    vertical='top', wrap_text=True)
            lines = max(len(key_facts),
                        (len(str(values[3])) // 50) + 1,
                        (len(str(values[4])) // 45) + 1)
            ws.row_dimensions[row].height = max(40, min(lines * 16, 220))
            row += 1

    def _collect_footnotes(self):
        """Collect all footnotes from the analysis.

        First tries to use the compiled footnotes array from the result (contains all citations).
        Falls back to collecting individual question footnotes if compiled array not available.
        """
        footnotes = []

        # Try compiled footnotes array first (contains ALL citations from document)
        compiled_footnotes = self.result.get('footnotes', [])
        if compiled_footnotes:
            for fn_num, fn_text in enumerate(compiled_footnotes, 1):
                if fn_text and fn_text.strip():
                    footnotes.append({
                        'number': fn_num,
                        'section': '',  # Compiled footnotes don't have section context
                        'question': '',
                        'footnote': fn_text.strip(),
                        'pages': []
                    })
            return footnotes

        # Fallback: collect from individual questions
        fn_num = 1
        for section in self.result.get('sections', []):
            for q in section.get('questions', []):
                fn = q.get('footnote')
                if fn and fn.strip():
                    footnotes.append({
                        'number': fn_num,
                        'section': section.get('section_name', 'Unknown'),
                        'question': q.get('question', ''),
                        'footnote': fn.strip(),
                        'pages': q.get('page_citations', [])
                    })
                    fn_num += 1
        return footnotes

    def _format_api_key_requirements(self, api_key_requirements):
        """Format API-extracted key requirements for Excel display"""
        requirements = {}

        # Map API keys to display-friendly names
        key_name_mapping = {
            'project_name': 'Project Name',
            'owner': 'Owner/Agency',
            'engineer': 'Engineer',
            'location': 'Location',
            'bid_deadline': 'Bid Deadline',
            'completion_date': 'Timeline',
            'scope': 'Scope',
            'linear_feet': 'Scope',
            'payment_terms': 'Payment',
            'warranty': 'Warranty',
            'liquidated_damages': 'Liquidated Damages',
            'bid_bond': 'Bonding',
            'performance_bond': 'Bonding',
            'certifications': 'Certifications',
            'insurance': 'Insurance',
        }

        for key, value in api_key_requirements.items():
            if value and str(value).strip() and str(value).lower() not in ['not found', 'n/a', 'not specified']:
                # Get display name from mapping, or format the key nicely
                display_name = key_name_mapping.get(key, key.replace('_', ' ').title())

                # Clean the value - remove PDF citations for display
                clean_value = re.sub(r'<PDF pg[^>]+>', '', str(value)).strip()
                clean_value = re.sub(r'\s+', ' ', clean_value)

                # Truncate if too long
                if len(clean_value) > 200:
                    clean_value = clean_value[:197] + '...'

                # Merge with existing if same display name (e.g., multiple bond types)
                if display_name in requirements and clean_value not in requirements[display_name]:
                    requirements[display_name] = f"{requirements[display_name]}; {clean_value}"
                else:
                    requirements[display_name] = clean_value

        return requirements

    def _extract_key_requirements(self):
        """Extract key project requirements from analysis answers"""
        requirements = {}

        for section in self.result.get('sections', []):
            for q in section.get('questions', []):
                question_text = q.get('question', '').lower()
                answer = q.get('answer', '')

                if not answer or not answer.strip():
                    continue

                # Clean answer - remove PDF citations for display
                clean_answer = re.sub(r'<PDF pg[^>]+>', '', answer).strip()
                clean_answer = re.sub(r'\s+', ' ', clean_answer)

                # Truncate if too long
                if len(clean_answer) > 200:
                    clean_answer = clean_answer[:197] + '...'

                # Check against key requirement patterns
                for req_name, patterns in self.KEY_REQUIREMENT_PATTERNS.items():
                    if req_name not in requirements:
                        for pattern in patterns:
                            if pattern in question_text:
                                requirements[req_name] = clean_answer
                                break

        return requirements

    def _calculate_statistics(self):
        """Calculate overall statistics"""
        all_questions = []
        total_confidence = 0
        confidence_count = 0

        for section in self.result.get('sections', []):
            for q in section.get('questions', []):
                all_questions.append(q)
                # Try to get confidence if available
                conf = q.get('confidence', 0)
                if conf and conf > 0:
                    total_confidence += conf
                    confidence_count += 1

        total = len(all_questions)
        answered = sum(1 for q in all_questions if q.get('answer'))
        unanswered = total - answered
        answer_rate = (answered / total * 100) if total > 0 else 0
        avg_confidence = (total_confidence / confidence_count * 100) if confidence_count > 0 else 0

        return {
            'total': total,
            'answered': answered,
            'unanswered': unanswered,
            'answer_rate': answer_rate,
            'avg_confidence': avg_confidence
        }

    def _create_executive_summary(self):
        """Sheet 1: Executive Summary - Analysis statistics and key requirements"""
        ws = self.wb.create_sheet('Executive Summary', 0)
        stats = self._calculate_statistics()

        # === TITLE BANNER ===
        ws.merge_cells('A1:C2')
        title_cell = ws['A1']
        title_cell.value = 'BIDBRIEF ANALYSIS REPORT'
        title_cell.font = self.TITLE_FONT
        title_cell.fill = self.TITLE_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

        row = 4

        # === PARTIAL RESULTS BANNER ===
        if self.is_partial:
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = '⚠ PARTIAL RESULTS - Analysis was stopped before completion'
            ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="92400E")
            ws[f'A{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 30
            row += 2

        # === ANALYSIS STATISTICS SECTION ===
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = 'Analysis Statistics'
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 28
        row += 1

        # Statistics rows
        stat_data = [
            ('Total Questions:', str(stats['total'])),
            ('Questions Answered:', str(stats['answered'])),
            ('Answer Rate:', f"{stats['answer_rate']:.0f}%"),
            ('Average Confidence:', f"{stats['avg_confidence']:.0f}%" if stats['avg_confidence'] > 0 else 'N/A'),
            ('Analysis Date:', self._get_timestamp()),
        ]

        for idx, (label, value) in enumerate(stat_data):
            is_alt = idx % 2 == 1

            ws.cell(row, 1, label).font = self.STAT_LABEL_FONT
            ws.cell(row, 2, value).font = self.STAT_VALUE_FONT

            for col in range(1, 4):
                cell = ws.cell(row, col)
                cell.border = self.BORDER_THIN
                if is_alt:
                    cell.fill = self.ALT_ROW_FILL

            ws.row_dimensions[row].height = 22
            row += 1

        row += 1

        # === KEY DOCUMENT DETAILS SECTION ===
        if self.key_requirements:
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = 'Key Document Details'
            ws[f'A{row}'].font = self.SECTION_FONT
            ws[f'A{row}'].fill = self.SECTION_FILL
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 28
            row += 1

            # Sort requirements by a preferred order
            preferred_order = ['Project Name', 'Owner/Agency', 'Engineer', 'Location', 'Timeline',
                             'Scope', 'Bid Deadline', 'Payment', 'Warranty', 'Liquidated Damages',
                             'Bonding', 'Certifications', 'Insurance']

            sorted_reqs = []
            for key in preferred_order:
                if key in self.key_requirements:
                    sorted_reqs.append((key + ':', self.key_requirements[key]))

            # Add any remaining that weren't in preferred order
            for key, value in self.key_requirements.items():
                if key not in preferred_order:
                    sorted_reqs.append((key + ':', value))

            for idx, (label, value) in enumerate(sorted_reqs):
                is_alt = idx % 2 == 1

                ws.cell(row, 1, label).font = self.STAT_LABEL_FONT
                ws.merge_cells(f'B{row}:C{row}')
                ws.cell(row, 2, value).font = self.DATA_FONT
                ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')

                for col in range(1, 4):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if is_alt:
                        cell.fill = self.ALT_ROW_FILL

                # Calculate dynamic row height based on content length
                # Columns B+C combined width is ~70 chars, each line ~15px height
                chars_per_line = 60
                num_lines = max(1, (len(value) // chars_per_line) + 1)
                row_height = max(25, num_lines * 18)  # Minimum 25, 18px per line
                ws.row_dimensions[row].height = row_height
                row += 1
        else:
            # No key requirements found
            row += 1
            ws[f'A{row}'] = 'Key project requirements will appear here when extracted from the analysis.'
            ws[f'A{row}'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35

    def _create_detailed_results(self):
        """Sheet 2: Detailed Results - All Q&A in unified professional table"""
        ws = self.wb.create_sheet('Detailed Results')

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = 'COMPLETE ANALYSIS RESULTS'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="104090")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Headers — Answer Summary (L6.5 distilled answer) sits between the
        # appended-quotes Answer column and PDF Pages.
        row = 2
        headers = ['#', 'Section', 'Question', 'Answer', 'Answer Summary', 'PDF Pages', 'Footnote', 'Status']
        widths = [5, 25, 45, 60, 40, 12, 8, 11]

        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER_THIN
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[row].height = 25

        # Data rows
        row = 3
        question_num = 1
        current_section = None

        for section in self.result.get('sections', []):
            section_name = section.get('section_name', '')

            for q in section.get('questions', []):
                answer_text = sanitize_for_excel(q.get('answer') or '')
                has_answer = bool(answer_text and answer_text.strip())
                pages = q.get('page_citations', [])
                pages_str = ', '.join(map(str, pages)) if pages else '-'
                has_footnote = bool(q.get('footnote'))

                is_section_start = section_name != current_section
                is_alt_row = question_num % 2 == 0

                ws.cell(row, 1, question_num).font = self.DATA_FONT
                ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

                section_cell = ws.cell(row, 2, section_name if is_section_start else '')
                if is_section_start:
                    section_cell.font = self.DATA_FONT_BOLD
                    current_section = section_name
                else:
                    section_cell.font = self.DATA_FONT

                ws.cell(row, 3, sanitize_for_excel(q.get('question', ''))).font = self.DATA_FONT

                answer_cell = ws.cell(row, 4, answer_text if has_answer else 'Not found in document')
                if has_answer:
                    answer_cell.font = self.DATA_FONT
                else:
                    answer_cell.font = Font(name='Calibri', size=11, italic=True, color="9CA3AF")

                summary_text = sanitize_for_excel(q.get('answer_summary') or '')
                summary_cell = ws.cell(row, 5, summary_text if summary_text else '-')
                summary_cell.font = self.DATA_FONT if summary_text else Font(
                    name='Calibri', size=11, italic=True, color="9CA3AF")

                pages_cell = ws.cell(row, 6, pages_str)
                pages_cell.font = self.DATA_FONT_BOLD if has_answer else self.DATA_FONT
                pages_cell.alignment = Alignment(horizontal='center', vertical='top')

                fn_cell = ws.cell(row, 7, '✓' if has_footnote else '-')
                fn_cell.font = Font(name='Calibri', size=11, bold=True, color=self.GREEN if has_footnote else self.GRAY)
                fn_cell.alignment = Alignment(horizontal='center', vertical='top')

                status_cell = ws.cell(row, 8, 'Found' if has_answer else 'Not Found')
                status_cell.font = Font(name='Calibri', size=10, bold=True)
                status_cell.fill = self.ANSWERED_FILL if has_answer else self.UNANSWERED_FILL
                status_cell.alignment = Alignment(horizontal='center', vertical='top')

                for col in range(1, 9):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if col not in [7, 8]:
                        if is_alt_row:
                            cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(
                        horizontal='left' if col in [2, 3, 4, 5] else 'center',
                        vertical='top',
                        wrap_text=True
                    )

                ws.row_dimensions[row].height = 55
                row += 1
                question_num += 1

    def _create_by_section(self):
        """Sheet 3: By Section - Questions grouped with section breakdown"""
        ws = self.wb.create_sheet('By Section')

        row = 1
        for section_idx, section in enumerate(self.result.get('sections', [])):
            questions = section.get('questions', [])
            answered = sum(1 for q in questions if q.get('answer') and q.get('answer').strip())
            total = len(questions)
            rate = (answered / total * 100) if total > 0 else 0

            # Section header with stats
            ws.merge_cells(f'A{row}:G{row}')
            header_cell = ws[f'A{row}']
            header_text = f"{section.get('section_name', 'Unknown Section').upper()}  ({answered}/{total} answered - {rate:.0f}%)"
            header_cell.value = header_text
            header_cell.font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
            header_cell.fill = self.HEADER_FILL
            header_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 32
            row += 1

            # Column headers
            headers = ['#', 'Question', 'Answer', 'Answer Summary', 'PDF Pages', 'Notes', 'Status']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.font = self.SUBHEADER_FONT
                cell.fill = self.SUBHEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.BORDER_THIN
            ws.row_dimensions[row].height = 22
            row += 1

            # Questions
            for idx, q in enumerate(questions, start=1):
                answer_text = sanitize_for_excel(q.get('answer') or '')
                has_answer = bool(answer_text and answer_text.strip())
                pages = q.get('page_citations', [])
                pages_str = ', '.join(map(str, pages)) if pages else '-'
                is_alt = idx % 2 == 0

                ws.cell(row, 1, idx).font = self.DATA_FONT
                ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

                ws.cell(row, 2, sanitize_for_excel(q.get('question', ''))).font = self.DATA_FONT

                answer_cell = ws.cell(row, 3, answer_text if has_answer else 'Not found in document')
                if has_answer:
                    answer_cell.font = self.DATA_FONT
                else:
                    answer_cell.font = Font(name='Calibri', size=11, italic=True, color="9CA3AF")

                summary_text = sanitize_for_excel(q.get('answer_summary') or '')
                summary_cell = ws.cell(row, 4, summary_text if summary_text else '-')
                summary_cell.font = self.DATA_FONT if summary_text else Font(
                    name='Calibri', size=11, italic=True, color="9CA3AF")

                ws.cell(row, 5, pages_str).font = self.DATA_FONT_BOLD if has_answer else self.DATA_FONT
                ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='top')

                has_footnote = bool(q.get('footnote'))
                notes_cell = ws.cell(row, 6, 'See footnotes' if has_footnote else '-')
                notes_cell.font = Font(name='Calibri', size=10, italic=True, color=self.BLUE if has_footnote else self.GRAY)
                notes_cell.alignment = Alignment(horizontal='center', vertical='top')

                status_cell = ws.cell(row, 7, '✓' if has_answer else '✗')
                status_cell.font = Font(name='Calibri', size=14, bold=True, color=self.GREEN if has_answer else self.RED)
                status_cell.fill = self.ANSWERED_FILL if has_answer else self.UNANSWERED_FILL
                status_cell.alignment = Alignment(horizontal='center', vertical='top')

                for col in range(1, 8):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if col != 7 and is_alt:
                        cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(
                        horizontal='left' if col in [2, 3, 4] else 'center',
                        vertical='top',
                        wrap_text=True
                    )

                ws.row_dimensions[row].height = 50
                row += 1

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10

    def _create_footnotes_sheet(self):
        """Sheet 4: Footnotes - All citations and contextual notes"""
        ws = self.wb.create_sheet('Footnotes')

        # Title with styling
        ws.merge_cells('A1:B1')
        ws['A1'] = 'FOOTNOTES & CITATIONS'
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = self.TITLE_FILL
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 35

        if not self.footnotes:
            ws['A3'] = 'No footnotes were generated during this analysis.'
            ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            ws.column_dimensions['A'].width = 60
            return

        # Subtitle with count
        ws['A2'] = f'{len(self.footnotes)} footnotes/citations collected from analysis'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
        ws.row_dimensions[2].height = 25

        # Check if we have compiled footnotes (simple format) or question-based footnotes
        has_question_context = any(fn.get('section') or fn.get('question') for fn in self.footnotes)

        if has_question_context:
            # Format with section/question context
            row = 4
            headers = ['#', 'Section', 'Related Question', 'Footnote/Citation', 'PDF Pages']
            widths = [6, 25, 40, 55, 12]

            for col, (header, width) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.BORDER_THIN
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.row_dimensions[row].height = 25

            for fn in self.footnotes:
                row += 1
                is_alt = fn['number'] % 2 == 0

                ws.cell(row, 1, fn['number']).font = self.DATA_FONT_BOLD
                ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row, 1).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

                ws.cell(row, 2, sanitize_for_excel(fn['section'])).font = self.DATA_FONT
                question_text = sanitize_for_excel(fn['question'])
                ws.cell(row, 3, question_text[:100] + '...' if len(question_text) > 100 else question_text).font = self.DATA_FONT
                ws.cell(row, 4, sanitize_for_excel(fn['footnote'])).font = self.DATA_FONT

                pages_str = ', '.join(map(str, fn['pages'])) if fn['pages'] else '-'
                ws.cell(row, 5, pages_str).font = self.DATA_FONT_BOLD
                ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='top')

                for col in range(1, 6):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if is_alt and col != 1:  # Don't override # column highlight
                        cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(
                        horizontal='left' if col in [2, 3, 4] else 'center',
                        vertical='top',
                        wrap_text=True
                    )

                ws.row_dimensions[row].height = 45
        else:
            # Simple format - just # and footnote text (for compiled footnotes)
            row = 4
            headers = ['#', 'Footnote/Citation']
            widths = [6, 100]

            for col, (header, width) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center' if col == 1 else 'left', vertical='center', wrap_text=True)
                cell.border = self.BORDER_THIN
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.row_dimensions[row].height = 25

            for fn in self.footnotes:
                row += 1
                is_alt = fn['number'] % 2 == 0

                # Number column with highlight
                cell_num = ws.cell(row, 1, fn['number'])
                cell_num.font = self.DATA_FONT_BOLD
                cell_num.alignment = Alignment(horizontal='center', vertical='top')
                cell_num.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                cell_num.border = self.BORDER_THIN

                # Footnote text
                cell_fn = ws.cell(row, 2, sanitize_for_excel(fn['footnote']))
                cell_fn.font = self.DATA_FONT
                cell_fn.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell_fn.border = self.BORDER_THIN
                if is_alt:
                    cell_fn.fill = self.ALT_ROW_FILL

                ws.row_dimensions[row].height = 50

    def _create_optimized_scan_sheet(self):
        """Sheet: Optimized Scan - V2 Pipeline Document Navigator Audit Data

        Shows pages targeted by each expert and keywords searched for audit purposes.
        Only created when optimized_scan_data is available (V2 pipeline analyses).
        """
        ws = self.wb.create_sheet('Optimized Scan')

        # Check if we have data
        if not self.optimized_scan_data or len(self.optimized_scan_data) == 0:
            ws.merge_cells('A1:C1')
            ws['A1'] = 'STAGE NOT PROCESSED BY USER'
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="9CA3AF")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40
            ws.merge_cells('A3:C3')
            ws['A3'] = 'Optimized Scan Pass was not run for this analysis'
            ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            ws.column_dimensions['A'].width = 50
            return

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'V2 PIPELINE - OPTIMIZED SCAN AUDIT'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="104090")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle explaining what this is
        ws.merge_cells('A2:F2')
        ws['A2'] = 'Document Navigator pre-scan analysis showing which pages each expert was directed to and why'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
        ws.row_dimensions[2].height = 25

        row = 4

        # Document Structure Summary
        structure = self.optimized_scan_data.get('structure', {})
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'DOCUMENT STRUCTURE DETECTED'
        ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = self.HEADER_FILL
        ws.row_dimensions[row].height = 28
        row += 1

        structure_items = [
            ('Table of Contents', 'Yes' if structure.get('has_toc') else 'No', f"{structure.get('toc_entries_count', 0)} entries"),
            ('Index', 'Yes' if structure.get('has_index') else 'No', f"{structure.get('index_terms_count', 0)} terms"),
            ('Appendix', 'Yes' if structure.get('has_appendix') else 'No', f"{len(structure.get('appendix_pages', []))} pages"),
        ]

        for label, found, details in structure_items:
            ws.cell(row, 1, label).font = self.STAT_LABEL_FONT
            ws.cell(row, 2, found).font = self.DATA_FONT_BOLD
            ws.cell(row, 2).fill = self.ANSWERED_FILL if found == 'Yes' else self.UNANSWERED_FILL
            ws.cell(row, 3, details).font = self.DATA_FONT
            for col in range(1, 4):
                ws.cell(row, col).border = self.BORDER_THIN
            row += 1

        row += 1  # Spacer

        # Expert Assignments
        expert_assignments = self.optimized_scan_data.get('expert_assignments', [])
        all_keywords = self.optimized_scan_data.get('all_keywords_by_section', {})

        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'EXPERT PAGE ASSIGNMENTS'
        ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = self.HEADER_FILL
        ws.row_dimensions[row].height = 28
        row += 1

        # Headers for expert assignments
        headers = ['Expert', 'Section', 'Primary Pages', 'Context Pages', 'Total Pages', 'Keywords Matched']
        widths = [25, 20, 25, 25, 12, 50]

        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.SUBHEADER_FONT
            cell.fill = self.SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER_THIN
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[row].height = 25
        row += 1

        if expert_assignments:
            for idx, assignment in enumerate(expert_assignments):
                is_alt = idx % 2 == 1

                primary_pages = assignment.get('primary_pages', [])
                context_pages = assignment.get('context_pages', [])
                keywords_matched = assignment.get('keywords_matched', [])

                ws.cell(row, 1, assignment.get('expert', '')).font = self.DATA_FONT_BOLD
                ws.cell(row, 2, assignment.get('section_id', '')).font = self.DATA_FONT
                ws.cell(row, 3, ', '.join(map(str, primary_pages[:15])) + ('...' if len(primary_pages) > 15 else '')).font = self.DATA_FONT
                ws.cell(row, 4, ', '.join(map(str, context_pages[:15])) + ('...' if len(context_pages) > 15 else '')).font = self.DATA_FONT
                ws.cell(row, 5, assignment.get('total_pages', len(primary_pages) + len(context_pages))).font = self.DATA_FONT_BOLD
                ws.cell(row, 6, ', '.join(keywords_matched[:5]) + ('...' if len(keywords_matched) > 5 else '')).font = self.DATA_FONT

                for col in range(1, 7):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if is_alt:
                        cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(horizontal='left' if col in [1, 2, 3, 4, 6] else 'center', vertical='top', wrap_text=True)

                ws.row_dimensions[row].height = 35
                row += 1
        else:
            ws.cell(row, 1, 'No expert assignments (document structure not found or all questions require exhaustive scan)')
            ws.cell(row, 1).font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            row += 1

        row += 1  # Spacer

        # Keywords by Section
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'ALL KEYWORDS SEARCHED BY SECTION'
        ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = self.HEADER_FILL
        ws.row_dimensions[row].height = 28
        row += 1

        headers = ['Section ID', 'Keywords Searched']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.SUBHEADER_FONT
            cell.fill = self.SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_THIN
        ws.row_dimensions[row].height = 25
        row += 1

        if all_keywords:
            for idx, (section_id, keywords) in enumerate(all_keywords.items()):
                is_alt = idx % 2 == 1

                ws.cell(row, 1, section_id).font = self.DATA_FONT_BOLD
                ws.cell(row, 2, ', '.join(keywords) if keywords else 'No keywords extracted').font = self.DATA_FONT

                for col in range(1, 3):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if is_alt:
                        cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                ws.row_dimensions[row].height = max(25, min(80, len(keywords) * 2))  # Dynamic height
                row += 1
        else:
            ws.cell(row, 1, 'No keywords data available')
            ws.cell(row, 1).font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            row += 1

        row += 1  # Spacer

        # Reduction Estimate
        reduction = self.optimized_scan_data.get('estimated_reduction', 'N/A')
        unassigned = self.optimized_scan_data.get('unassigned_questions', [])

        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'EFFICIENCY SUMMARY'
        ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = self.HEADER_FILL
        ws.row_dimensions[row].height = 28
        row += 1

        ws.cell(row, 1, 'Estimated Page Reduction:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, reduction).font = self.DATA_FONT_BOLD
        ws.cell(row, 2).fill = self.ANSWERED_FILL
        row += 1

        ws.cell(row, 1, 'Questions Requiring Full Scan:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, str(len(unassigned)) if unassigned else '0').font = self.DATA_FONT_BOLD
        row += 1

    def _create_unanswered_pass_sheet(self):
        """Sheet: Unanswered Questions Pass - V2 Pipeline second pass data"""
        ws = self.wb.create_sheet('Unanswered Pass')

        # Check if we have data
        if not self.unanswered_pass_data or len(self.unanswered_pass_data) == 0:
            ws.merge_cells('A1:C1')
            ws['A1'] = 'STAGE NOT PROCESSED BY USER'
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="9CA3AF")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40
            ws.merge_cells('A3:C3')
            ws['A3'] = 'Unanswered Questions Pass was not run for this analysis'
            ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            ws.column_dimensions['A'].width = 50
            return

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = 'V2 PIPELINE - UNANSWERED QUESTIONS PASS'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="104090")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells('A2:D2')
        ws['A2'] = 'Second pass focused on questions that remained unanswered after exhaustive scan'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
        ws.row_dimensions[2].height = 25

        row = 4
        questions_processed = self.unanswered_pass_data.get('questions_processed', [])
        answers_found = self.unanswered_pass_data.get('answers_found', 0)
        total_questions = self.unanswered_pass_data.get('total_questions', len(questions_processed))

        # Summary stats
        ws.cell(row, 1, 'Questions Processed:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, str(total_questions)).font = self.DATA_FONT_BOLD
        row += 1
        ws.cell(row, 1, 'New Answers Found:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, str(answers_found)).font = self.DATA_FONT_BOLD
        ws.cell(row, 2).fill = self.ANSWERED_FILL
        row += 2

        if questions_processed:
            # Headers
            headers = ['Question', 'Section', 'Found Answer']
            widths = [50, 20, 15]

            for col, (header, width) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.BORDER_THIN
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.row_dimensions[row].height = 25
            row += 1

            for idx, q in enumerate(questions_processed):
                is_alt = idx % 2 == 1
                found = q.get('found', False)

                ws.cell(row, 1, sanitize_for_excel(q.get('question', 'Unknown'))).font = self.DATA_FONT
                ws.cell(row, 2, q.get('section', '')).font = self.DATA_FONT
                status_cell = ws.cell(row, 3, 'Yes' if found else 'No')
                status_cell.font = Font(name='Calibri', size=10, bold=True)
                status_cell.fill = self.ANSWERED_FILL if found else self.UNANSWERED_FILL
                status_cell.alignment = Alignment(horizontal='center', vertical='center')

                for col in range(1, 4):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if col != 3 and is_alt:
                        cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(horizontal='left' if col in [1, 2] else 'center', vertical='top', wrap_text=True)

                ws.row_dimensions[row].height = 35
                row += 1

    def _create_rag_sheet(self):
        """Sheet: Deep RAG - V2 Pipeline RAG analysis data"""
        ws = self.wb.create_sheet('Deep RAG')

        # Check if we have data
        if not self.rag_data or len(self.rag_data) == 0:
            ws.merge_cells('A1:C1')
            ws['A1'] = 'STAGE NOT PROCESSED BY USER'
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="9CA3AF")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40
            ws.merge_cells('A3:C3')
            ws['A3'] = 'Deep RAG analysis was not run for this analysis'
            ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            ws.column_dimensions['A'].width = 50
            return

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = 'V2 PIPELINE - DEEP RAG ANALYSIS'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="104090")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells('A2:E2')
        ws['A2'] = 'Final pass using Retrieval-Augmented Generation for remaining unanswered questions'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
        ws.row_dimensions[2].height = 25

        row = 4
        questions_processed = self.rag_data.get('questions_processed', [])
        answers_found = self.rag_data.get('answers_found', 0)
        total_questions = self.rag_data.get('total_questions', len(questions_processed))
        chunks_searched = self.rag_data.get('chunks_searched', 0)

        # Summary stats
        ws.cell(row, 1, 'Questions Processed:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, str(total_questions)).font = self.DATA_FONT_BOLD
        row += 1
        ws.cell(row, 1, 'New Answers Found:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, str(answers_found)).font = self.DATA_FONT_BOLD
        ws.cell(row, 2).fill = self.ANSWERED_FILL
        row += 1
        ws.cell(row, 1, 'Chunks Searched:').font = self.STAT_LABEL_FONT
        ws.cell(row, 2, str(chunks_searched)).font = self.DATA_FONT_BOLD
        row += 2

        if questions_processed:
            # Headers
            headers = ['Question', 'Section', 'Similarity', 'Found Answer']
            widths = [45, 20, 12, 15]

            for col, (header, width) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.BORDER_THIN
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.row_dimensions[row].height = 25
            row += 1

            for idx, q in enumerate(questions_processed):
                is_alt = idx % 2 == 1
                found = q.get('found', False)
                similarity = q.get('similarity', 0)
                similarity_str = f"{similarity * 100:.1f}%" if similarity else 'N/A'

                ws.cell(row, 1, sanitize_for_excel(q.get('question', 'Unknown'))).font = self.DATA_FONT
                ws.cell(row, 2, q.get('section', '')).font = self.DATA_FONT
                ws.cell(row, 3, similarity_str).font = self.DATA_FONT
                ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')
                status_cell = ws.cell(row, 4, 'Yes' if found else 'No')
                status_cell.font = Font(name='Calibri', size=10, bold=True)
                status_cell.fill = self.ANSWERED_FILL if found else self.UNANSWERED_FILL
                status_cell.alignment = Alignment(horizontal='center', vertical='center')

                for col in range(1, 5):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if col not in [3, 4] and is_alt:
                        cell.fill = self.ALT_ROW_FILL
                    if col in [1, 2]:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                ws.row_dimensions[row].height = 35
                row += 1

    def _get_timestamp(self):
        """Get formatted timestamp"""
        return datetime.now().strftime('%B %d, %Y')
